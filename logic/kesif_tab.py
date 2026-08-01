import streamlit as st
import pandas as pd
import sqlite3
from datetime import datetime
import os
import re
import traceback
import pdfplumber
import io
import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font

from logic.db_utils import get_connection, get_settings_connection

POZ_PATTERN = re.compile(
    r'(?<![A-Z0-9])(?:[A-ZÇĞİÖŞÜ]{1,8}-\d{1,6}(?![./-]\d)|(?:[A-ZÇĞİÖŞÜ]{1,8}[./-]?)?\d{1,5}(?:[./-]\d{1,6})+)(?![A-Z0-9])',
    re.IGNORECASE
)
PRICE_PATTERN = re.compile(r'(?<!\d)(?:\d{1,3}(?:[.\s]\d{3})+|\d+)(?:,\d{1,2}|\.\d{1,2})(?!\d)')


def normalize_poz_no(value):
    text = str(value or '').upper().strip()
    text = text.replace('–', '-').replace('—', '-').replace(' ', '')
    return text.replace('İ', 'I')


def clean_price_value(value):
    text = str(value or '').upper().replace('TL', '').replace('TRY', '').replace('₺', '').strip()
    text = text.replace(' ', '')
    if not text:
        return 0.0
    try:
        if ',' in text and '.' in text:
            text = text.replace('.', '').replace(',', '.')
        elif ',' in text:
            text = text.replace(',', '.')
        return float(re.search(r'[-+]?\d*\.?\d+', text).group())
    except (AttributeError, ValueError):
        return 0.0


def extract_price_rows_from_text(text):
    """Extract standard unit-price rows from imperfect PDF text output."""
    rows = []
    known_units = r'\b(ADET|M2|M3|M|KG|TON|SAAT|GÜN|GUN|KM|TAKIM|SET|LT|LİTRE)\b'
    for raw_line in str(text or '').splitlines():
        line = ' '.join(raw_line.split())
        poz_match = POZ_PATTERN.search(line)
        price_matches = list(PRICE_PATTERN.finditer(line))
        if not poz_match or not price_matches:
            continue
        price_match = price_matches[-1]
        poz_no = normalize_poz_no(poz_match.group())
        price = clean_price_value(price_match.group())
        if not poz_no or price <= 0:
            continue
        unit_match = re.search(known_units, line, flags=re.IGNORECASE)
        unit = unit_match.group().upper() if unit_match else 'Adet'
        description = (line[:poz_match.start()] + ' ' + line[poz_match.end():price_match.start()]).strip(' -:')
        rows.append([poz_no, description or '-', unit, price])
    return rows


def is_valid_poz_no(value):
    poz_no = normalize_poz_no(value)
    return bool(POZ_PATTERN.fullmatch(poz_no))


# Re-export if needed, but better to use db_utils

@st.cache_resource
def ensure_discovery_tables():
    """Create all required Keşif and Protocol tables in the settings database.
    Also migrates data from main DB if this is a first-time setup or database swap.
    """
    conn_data = get_connection()
    conn_sett = get_settings_connection()
    
    cursor_sett = conn_sett.cursor()
    cursor_data = conn_data.cursor()

    # Tables to migrate if they exist in Data DB but not in Settings DB
    tables_to_migrate = [
        "users", "lab_staff", "test_durations", 
        "kesif_birim_fiyatlar", "kesif_ayarlari", "kesif_ek_kalemler", 
        "kesif_kayitlari", "kesif_akredite_durumlari", "word_protokoller", "protocols"
    ]

    for table in tables_to_migrate:
        # Check if table exists in Data DB
        cursor_data.execute(f"SELECT name FROM sqlite_master WHERE type='table' AND name='{table}'")
        exists_in_data = cursor_data.fetchone()
        
        # Check if table exists in Settings DB
        cursor_sett.execute(f"SELECT name FROM sqlite_master WHERE type='table' AND name='{table}'")
        exists_in_sett = cursor_sett.fetchone()

        # If it exists in Data but not in Settings, migrate it
        if exists_in_data and not exists_in_sett:
            try:
                # Get schema from Data DB
                cursor_data.execute(f"SELECT sql FROM sqlite_master WHERE type='table' AND name='{table}'")
                schema_sql = cursor_data.fetchone()[0]
                
                # Create in Settings DB
                cursor_sett.execute(schema_sql)
                
                # Copy data
                df = pd.read_sql(f"SELECT * FROM {table}", conn_data)
                if not df.empty:
                    df.to_sql(table, conn_sett, if_exists='append', index=False)
                    print(f"Migrated table {table} ({len(df)} rows) to settings database.")
            except Exception as e:
                print(f"Error migrating {table}: {e}")

    # Now define/update all tables in Settings DB (to ensure latest schema)
    
    # 1. word_protokoller
    cursor_sett.execute("""
        CREATE TABLE IF NOT EXISTS word_protokoller (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_name TEXT, firm TEXT, tax_id TEXT, protocol_date TEXT,
            base_cost REAL, kdv_amount REAL, stamp_tax REAL, turkak_fee REAL, total_amount REAL,
            created_date TEXT, file_blob BLOB
        )
    """)
    
    # 2. kesif_birim_fiyatlar
    cursor_sett.execute("""
        CREATE TABLE IF NOT EXISTS kesif_birim_fiyatlar (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            poz_no TEXT, description TEXT, unit TEXT, price REAL, year INTEGER, lab_type TEXT
        )
    """)

    # 3. kesif_ayarlari
    cursor_sett.execute("""
        CREATE TABLE IF NOT EXISTS kesif_ayarlari (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_name TEXT UNIQUE,
            protocol_rate REAL DEFAULT 10.0, acc_rate REAL DEFAULT 10.0, vat_rate REAL DEFAULT 20.0,
            year INTEGER DEFAULT 2025
        )
    """)
    
    # helper for columns
    def add_col(db_cursor, table, col, def_type):
        try:
            db_cursor.execute(f"PRAGMA table_info({table})")
            cols = [c[1].lower() for c in db_cursor.fetchall()]
            if col.lower() not in cols:
                db_cursor.execute(f"ALTER TABLE {table} ADD COLUMN {col} {def_type}")
        except Exception as e:
            print(f"Schema update error ({table}.{col}): {e}")

    for col in ['formula_c', 'formula_d', 'formula_e', 'formula_f', 'formula_g', 'stamp_tax_rate', 'project_coeff_rate', 'turkak_rate']:
        add_col(cursor_sett, 'kesif_ayarlari', col, 'TEXT')

    # 4. kesif_ek_kalemler
    cursor_sett.execute("""
        CREATE TABLE IF NOT EXISTS kesif_ek_kalemler (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_name TEXT, poz_no TEXT, description TEXT, unit TEXT,
            price REAL, quantity REAL, order_index INTEGER, is_accredited INTEGER DEFAULT 0
        )
    """)

    # 5. kesif_kayitlari
    cursor_sett.execute("""
        CREATE TABLE IF NOT EXISTS kesif_kayitlari (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_name TEXT, year INTEGER, total_amount REAL, created_at TEXT,
            status TEXT DEFAULT 'BEKLİYOR', dekont_no TEXT,
            payment_date TEXT, bank_info TEXT
        )
    """)
    add_col(cursor_sett, 'kesif_kayitlari', 'payment_date', 'TEXT')
    add_col(cursor_sett, 'kesif_kayitlari', 'bank_info', 'TEXT')

    # 6. kesif_akredite_durumlari
    cursor_sett.execute("""
        CREATE TABLE IF NOT EXISTS kesif_akredite_durumlari (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_name TEXT, poz_no TEXT, is_accredited INTEGER DEFAULT 0,
            UNIQUE(project_name, poz_no)
        )
    """)

    # 7. protocols
    cursor_sett.execute("""
        CREATE TABLE IF NOT EXISTS protocols (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sequence_no INTEGER, office_record_no TEXT, protocol_no TEXT, 
            region_no INTEGER, sender TEXT, firm TEXT, job_description TEXT, 
            protocol_date TEXT, base_cost REAL, kdv_amount REAL, 
            total_cost_with_kdv REAL, turkak_fee REAL, 
            secondary_keşif_with_kdv REAL, stamp_tax REAL, 
            total_amount REAL, payment_date TEXT, receipt_no TEXT, 
            bank_info TEXT, month TEXT, is_archived INTEGER DEFAULT 0, 
            archive_year TEXT, linked_project_name TEXT
        )
    """)

    # 8. test_durations (System)
    cursor_sett.execute("CREATE TABLE IF NOT EXISTS test_durations (id INTEGER PRIMARY KEY AUTOINCREMENT, test_name TEXT, lab_type TEXT)")
    add_col(cursor_sett, 'test_durations', 'poz_no', 'TEXT')
    add_col(cursor_sett, 'test_durations', 'is_accredited', 'INTEGER DEFAULT 0')
    add_col(cursor_sett, 'test_durations', 'duration_days', 'INTEGER DEFAULT 7')

    # 9. users (System)
    cursor_sett.execute("CREATE TABLE IF NOT EXISTS users (id INTEGER PRIMARY KEY AUTOINCREMENT, username TEXT UNIQUE, password TEXT)")
    add_col(cursor_sett, 'users', 'role', "TEXT DEFAULT 'user'")
    add_col(cursor_sett, 'users', 'permissions', "TEXT")
    add_col(cursor_sett, 'users', 'full_name', "TEXT")
    add_col(cursor_sett, 'users', 'title', "TEXT")

    # 10. lab_staff (System)
    cursor_sett.execute("CREATE TABLE IF NOT EXISTS lab_staff (id INTEGER PRIMARY KEY AUTOINCREMENT, lab_type TEXT, staff_name TEXT)")

    # 11. Data-specific columns in main DB
    # samples
    cursor_data.execute("CREATE TABLE IF NOT EXISTS samples (id INTEGER PRIMARY KEY AUTOINCREMENT, kayit_no TEXT)")
    for c, t in [
        ('lab_type', 'TEXT'), ('gelis_tarihi', 'TEXT'), ('yer', 'TEXT'), ('cins', 'TEXT'), ('miktar', 'TEXT'), ('miktar_birimi', 'TEXT'),
        ('aciklama', 'TEXT'), ('teslim_alan', 'TEXT'), ('teslim_eden', 'TEXT'), ('gonderen', 'TEXT'),
        ('evrak_kayit_no', 'TEXT'), ('ebays_tarih', 'TEXT'), ('ebays_sayi', 'TEXT'), ('deney_adi', 'TEXT'),
        ('deney_sorumlusu', 'TEXT DEFAULT "-"'), ('test_durumu', 'TEXT DEFAULT "Beklemede"'),
        ('deney_sonuc', 'TEXT DEFAULT "-"'), ('kayit_zamani', 'TEXT'),
        ('test_baslangic', 'TEXT DEFAULT "-"'), ('test_bitis', 'TEXT DEFAULT "-"'),
        ('rapor_tarihi', 'TEXT DEFAULT "-"'), ('rapor_sayisi', 'TEXT DEFAULT "-"'),
        ('proje', 'TEXT'), ('firma', 'TEXT')
    ]:
        add_col(cursor_data, 'samples', c, t)

    conn_sett.commit()
    conn_data.commit()
    conn_sett.close()
    conn_data.close()

def ensure_word_protokol_table():
    # Deprecated: use ensure_discovery_tables instead
    ensure_discovery_tables()

def save_word_protokol(project_name, firm, tax_id, protocol_date, costs, file_bytes):
    """Save Word protocol to database"""
    conn = get_settings_connection()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO word_protokoller (
            project_name, firm, tax_id, protocol_date, 
            base_cost, kdv_amount, stamp_tax, turkak_fee, total_amount,
            created_date, file_blob
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        project_name, firm, tax_id, protocol_date,
        costs['base_cost'], costs['kdv_amount'], costs['stamp_tax'], costs['turkak_fee'], costs['total_amount'],
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        file_bytes
    ))
    conn.commit()
    conn.close()

def get_all_word_protokoller():
    """Fetch all saved Word protocols"""
    conn = get_settings_connection()
    df = pd.read_sql("SELECT * FROM word_protokoller ORDER BY created_date DESC", conn)
    conn.close()
    return df

def delete_word_protokol(selected_id):
    """Delete Word protocol record"""
    conn = get_settings_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM word_protokoller WHERE id = ?", (selected_id,))
    conn.commit()
    conn.close()

def get_word_protokol_file(p_id):
    """Fetch binary file of a Word protocol"""
    conn = get_settings_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT file_blob FROM word_protokoller WHERE id = ?", (p_id,))
    row = cursor.fetchone()
    conn.close()
    return row[0] if row else None

def update_word_protokol(p_id, project_name, firm, tax_id, protocol_date, total_amount):
    """Update existing Word protocol metadata"""
    conn = get_settings_connection()
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE word_protokoller 
        SET project_name = ?, firm = ?, tax_id = ?, protocol_date = ?, total_amount = ?
        WHERE id = ?
    """, (project_name, firm, tax_id, protocol_date, total_amount, p_id))
    conn.commit()
    conn.close()

def parse_formula_for_ui(formula, context):
    """
    Safely evaluates an Excel-style formula for the UI display.
    formula: string like '=(P21+P22)*(1+{protocol_rate}/100)'
    context: dict mapping cell names (P21, P22...) and placeholders ({rate}) to values
    """
    if not formula or not isinstance(formula, str): return 0.0
    
    # 1. Clean the formula (remove '=' if exists)
    f = formula.strip()
    if f.startswith("="): f = f[1:]
    
    # 2. Replace placeholders with actual values
    for key, val in context.items():
        if key.startswith("{") and key.endswith("}"):
            f = f.replace(key, str(val))
        else:
            # For cell references like P21, we use a regex or simple replace if they are unique
            import re
            f = re.sub(rf'\b{key}\b', str(val), f)
            
    # 3. Simple safety check (allow only math chars, numbers and round)
    allowed_chars = "0123456789.+-*/() ,round"
    if any(c not in allowed_chars for c in f.lower()):
        return 0.0
        
    try:
        # Evaluate the math string with round() supported
        safe_dict = {"round": round, "abs": abs}
        return round(float(eval(f.lower(), {"__builtins__": {}}, safe_dict)), 2)
    except:
        return 0.0

def calculate_project_costs(project_name):
    """
    Calculates cost components (Base, KDV, Stamp, TURKAK, Total) for a given project
    based on the current Kesif configuration and data.
    """
    try:
        conn_data = get_connection()
        conn_sett = get_settings_connection()
        
        # 0. Defaults
        acc_rate, protocol_rate, project_coeff_rate = 10.0, 10.0, 0.0
        vat_rate, stamp_tax_rate, turkak_rate = 20.0, 0.00948, 1.0
        year = 2025
        
        def_c = "=(P21*(1+{acc_rate}/100)+P22)*(1+{protocol_rate}/100)*(1+{project_coeff_rate}/100)"
        def_d = "=ROUND(P23*{vat_rate}/100, 2)"
        def_e = "=ROUND(P23*{stamp_tax_rate}, 2)"
        def_f = "=ROUND(P21*{turkak_rate}/100, 2)"
        def_g = "=P23+P24+P25+P26"
        
        formula_c, formula_d, formula_e, formula_f, formula_g = def_c, def_d, def_e, def_f, def_g

        # 1. Resolve Project Name if Kayıt No is provided
        actual_project_name = project_name
        if str(project_name).isdigit() or (isinstance(project_name, str) and re.match(r'^\d+$', project_name)):
            df_res = pd.read_sql(f"SELECT proje FROM samples WHERE kayit_no='{project_name}' LIMIT 1", conn_data)
            if not df_res.empty:
                actual_project_name = df_res.iloc[0]['proje']
        
        # 2. Load Settings from Settings DB
        proj_settings = pd.read_sql(f"SELECT * FROM kesif_ayarlari WHERE project_name='{actual_project_name}'", conn_sett)
        
        if proj_settings.empty and actual_project_name != project_name:
            # Try if settings were saved under record number
            proj_settings = pd.read_sql(f"SELECT * FROM kesif_ayarlari WHERE project_name='{project_name}'", conn_sett)
            
        if not proj_settings.empty:
            s = proj_settings.iloc[0]
            year = int(s['year'])
            protocol_rate = float(s['protocol_rate'])
            acc_rate = float(s['acc_rate'])
            project_coeff_rate = float(s['project_coeff_rate']) if 'project_coeff_rate' in proj_settings.columns else 0.0
            vat_rate = float(s['vat_rate'])
            stamp_tax_rate = float(s['stamp_tax_rate']) if 'stamp_tax_rate' in proj_settings.columns else 0.00948
            turkak_rate = float(s['turkak_rate']) if 'turkak_rate' in proj_settings.columns else 1.0
            
            formula_c = s['formula_c'] if s.get('formula_c') else def_c
            formula_d = s['formula_d'] if s.get('formula_d') else def_d
            formula_e = s['formula_e'] if s.get('formula_e') else def_e
            formula_f = s['formula_f'] if s.get('formula_f') else def_f
            formula_g = s['formula_g'] if s.get('formula_g') else def_g

        # 3. Fetch Samples from Data DB
        if str(project_name).isdigit() or (isinstance(project_name, str) and re.match(r'^\d+$', project_name)):
            query_samples = f"SELECT deney_adi FROM samples WHERE kayit_no='{project_name}'"
        else:
            query_samples = f"SELECT deney_adi FROM samples WHERE proje='{project_name}'"
        
        df_samples = pd.read_sql(query_samples, conn_data)
        
        # 3. Calculate Items
        # Price Data from Settings DB
        df_prices = pd.read_sql(f"SELECT * FROM kesif_birim_fiyatlar WHERE year={year}", conn_sett)
        df_test_defs = pd.read_sql("SELECT test_name, poz_no, is_accredited FROM test_durations WHERE poz_no IS NOT NULL AND poz_no != ''", conn_sett)
        df_acc_overrides = pd.read_sql(f"SELECT poz_no, is_accredited FROM kesif_akredite_durumlari WHERE project_name='{actual_project_name}'", conn_sett)
        
        from collections import Counter
        # Robust normalization for matching
        df_samples['norm_name'] = df_samples['deney_adi'].str.strip().str.lower()
        exp_counts = Counter(df_samples['norm_name'].tolist())
        
        # Build a robust mapping from settings
        df_test_defs['norm_name'] = df_test_defs['test_name'].str.strip().str.lower()
        
        # FIX: Define acc_override_map from the fetched overrides
        acc_override_map = {r['poz_no'].strip(): bool(r['is_accredited']) for _, r in df_acc_overrides.iterrows()}
        
        robust_mapping = {}
        for _, r in df_test_defs.iterrows():
            p_no = r['poz_no'].strip()
            # Use override if exists, otherwise use default from test_durations
            is_acc = acc_override_map.get(p_no, bool(r['is_accredited']))
            robust_mapping[r['norm_name']] = (r['test_name'], p_no, is_acc)
        
        items = []
        for norm_name, count in exp_counts.items():
            if norm_name in robust_mapping:
                orig_name, poz, is_acc = robust_mapping[norm_name]
                # Flexible price lookup (TRIM)
                price_row = df_prices[df_prices['poz_no'].apply(normalize_poz_no) == normalize_poz_no(poz)]
                pr_val = 0.0
                if not price_row.empty:
                    pr_val = float(price_row.iloc[0]['price'])
                
                items.append({
                    'Tutar': float(count) * pr_val,
                    'is_accredited': is_acc
                })
                    
        # 4. Fetch Extra Items from Settings DB
        try:
            df_extras = pd.read_sql(f"SELECT price, quantity, is_accredited FROM kesif_ek_kalemler WHERE project_name='{actual_project_name}'", conn_sett)
        except:
            # Fallback if is_accredited is missing
            try:
                df_extras = pd.read_sql(f"SELECT price, quantity FROM kesif_ek_kalemler WHERE project_name='{actual_project_name}'", conn_sett)
                df_extras['is_accredited'] = 0
            except:
                df_extras = pd.DataFrame(columns=['price', 'quantity', 'is_accredited'])
                
        for _, r in df_extras.iterrows():
            items.append({
                'Tutar': float(r['price']) * float(r['quantity']),
                'is_accredited': bool(r.get('is_accredited', 0))
            })
            
        conn_data.close()
        conn_sett.close()
        # 5. Calculate Totals
        acc_total = sum(i['Tutar'] for i in items if i['is_accredited'])
        norm_total = sum(i['Tutar'] for i in items if not i['is_accredited'])
        
        # Context for formulas
        context = {
            "P21": acc_total, "P22": norm_total,
            "{acc_rate}": acc_rate, "{protocol_rate}": protocol_rate, "{project_coeff_rate}": project_coeff_rate,
            "{vat_rate}": vat_rate, "{stamp_tax_rate}": stamp_tax_rate, "{turkak_rate}": turkak_rate
        }
        
        c_val = parse_formula_for_ui(formula_c, context)
        context["P23"] = c_val
        d_val = parse_formula_for_ui(formula_d, context) # VAT
        context["P24"] = d_val
        e_val = parse_formula_for_ui(formula_e, context) # Stamp
        context["P25"] = e_val
        f_val = parse_formula_for_ui(formula_f, context) # TURKAK
        context["P26"] = f_val
        g_val = parse_formula_for_ui(formula_g, context) # Total
        
        totals = {
            "base_cost": c_val,
            "kdv_amount": d_val,
            "stamp_tax": e_val,
            "turkak_fee": f_val,
            "total_amount": g_val
        }
        return totals
        
    except Exception as e:
        import traceback
        print(f"ERROR calculating costs for {project_name}: {traceback.format_exc()}")
        return {
            "base_cost": 0.0, "kdv_amount": 0.0, "stamp_tax": 0.0, 
            "turkak_fee": 0.0, "total_amount": 0.0
        }

def render_kesif_tab():
    st.title("📑 Keşif ve Maliyet Analizi")
    
    # Ensure tables exist (Fix for swapped/missing database tables)
    ensure_discovery_tables()
    
    conn_sett = get_settings_connection()
    conn_data = get_connection()
    
    # 📄 Kayıtlı Word Protokoller Section
    with st.expander("📄 Kayıtlı Word Protokoller", expanded=False):
        df_word_protokols = get_all_word_protokoller()
        
        if not df_word_protokols.empty:
            st.dataframe(df_word_protokols[['id', 'project_name', 'firm', 'tax_id', 'protocol_date', 'total_amount', 'created_date']], use_container_width=True)
            
            # Action buttons
            col1, col2 = st.columns(2)
            with col1:
                selected_id = st.selectbox("Protokol Seç (ID)", df_word_protokols['id'].tolist())
            with col2:
                action = st.radio("İşlem", ["İndir", "Düzenle", "Sil"], horizontal=True)
            
            if action == "İndir":
                if st.button("📥 Word Dosyasını İndir"):
                    file_blob = get_word_protokol_file(selected_id)
                    if file_blob:
                        selected_row = df_word_protokols[df_word_protokols['id'] == selected_id].iloc[0]
                        st.download_button(
                            label="⬇️ İndir",
                            data=file_blob,
                            file_name=f"Protokol_{selected_row['project_name']}.docx",
                            mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document"
                        )
                    else:
                        st.error("Dosya bulunamadı!")
            
            elif action == "Düzenle":
                selected_row = df_word_protokols[df_word_protokols['id'] == selected_id].iloc[0]
                with st.form(f"edit_word_form_{selected_id}"):
                    st.markdown("#### Protokol Bilgilerini Düzenle")
                    e_col1, e_col2, e_col3 = st.columns(3)
                    e_firm = e_col1.text_input("Firma Adı", value=selected_row['firm'])
                    e_tax = e_col2.text_input("Vergi No", value=selected_row['tax_id'] if pd.notna(selected_row['tax_id']) else "")
                    e_date = e_col3.date_input("Protokol Tarihi", value=pd.to_datetime(selected_row['protocol_date']))
                    
                    e_col4, e_col5, e_col6, e_col7 = st.columns(4)
                    e_base = e_col4.number_input("Hizmet Bedeli", value=float(selected_row['base_cost']))
                    e_kdv = e_col5.number_input("KDV", value=float(selected_row['kdv_amount']))
                    e_stamp = e_col6.number_input("Damga", value=float(selected_row['stamp_tax']))
                    e_turkak = e_col7.number_input("TÜRKAK", value=float(selected_row['turkak_fee']))
                    
                    if st.form_submit_button("💾 Güncelle"):
                        costs = {
                            'base_cost': e_base,
                            'kdv_amount': e_kdv,
                            'stamp_tax': e_stamp,
                            'turkak_fee': e_turkak,
                            'total_amount': e_base + e_kdv + e_stamp + e_turkak
                        }
                        update_word_protokol(selected_id, e_firm, e_tax, e_date.strftime("%Y-%m-%d"), costs)
                        st.success("✅ Protokol güncellendi!")
                        st.rerun()
            
            elif action == "Sil":
                if st.button("🗑️ Protokolü Sil", type="secondary"):
                    delete_word_protokol(selected_id)
                    st.success("✅ Protokol silindi!")
                    st.rerun()
        else:
            st.info("Henüz kaydedilmiş Word Protokol bulunmuyor.")
    
    # 📌 Kayıtlı Keşifler ve Ödeme Takibi Section
    with st.expander("📌 Kayıtlı Keşifler ve Ödeme Takibi", expanded=False):
        # Add Search Bar
        search_history = st.text_input("🔍 Kayıtlı Keşiflerde Ara (Proje Adı)", placeholder="Proje isminden kelimeler girin...")
        
        query_saved = "SELECT * FROM kesif_kayitlari ORDER BY id DESC"
        df_saved = pd.read_sql(query_saved, conn_sett)
        
        if not df_saved.empty:
            # Filter by search term
            if search_history:
                df_saved = df_saved[df_saved['project_name'].str.contains(search_history, case=False, na=False)]
            
            st.info("Dekont No girdiğinizde durum otomatik olarak 'ÖDENDİ' olur.")
            
            # Prepare for display with a selection column
            display_saved = df_saved.copy()
            display_saved.insert(0, "Seç", False)
            
            # Fetch linked projects from protocols to show link status
            try:
                protocol_links = pd.read_sql("SELECT DISTINCT linked_project_name FROM protocols WHERE linked_project_name IS NOT NULL", conn_sett)['linked_project_name'].tolist()
                display_saved['Protokol'] = display_saved['project_name'].apply(lambda x: "🔗 AKTARILDI" if x in protocol_links else "⏳ BEKLEMEDE")
            except:
                display_saved['Protokol'] = "---"

            edited_saved = st.data_editor(
                display_saved, 
                key="kesif_history_editor",
                use_container_width=True,
                column_config={
                    "Seç": st.column_config.CheckboxColumn("Seç", default=False),
                    "id": None, # Hide ID
                    "project_name": "Proje Adı",
                    "year": "Yıl",
                    "total_amount": st.column_config.NumberColumn("Toplam Tutar (₺)", format="%.2f"),
                    "dekont_no": "Dekont No",
                    "payment_date": "Ödeme Tarihi",
                    "bank_info": "Banka / Açıklama",
                    "status": "Durum",
                    "created_at": "Kayıt Tarihi",
                    "Protokol": st.column_config.TextColumn("Protokol", help="Protokol listesine aktarılıp aktarılmadığını gösterir.")
                },
                disabled=["project_name", "year", "total_amount", "status", "created_at", "Protokol"]
            )
            
            # 1. Update edited changes (Dekont No, Payment Date, Bank Info)
            if not edited_saved.drop(columns=["Seç"]).equals(display_saved.drop(columns=["Seç"])):
                cursor = conn_sett.cursor()
                for i, row in edited_saved.iterrows():
                    old_row = display_saved.iloc[i]
                    if (row['dekont_no'] != old_row['dekont_no'] or 
                        row['payment_date'] != old_row['payment_date'] or 
                        row['bank_info'] != old_row['bank_info']):
                        
                        proj_name = row['project_name']
                        new_status = "✅ ÖDENDİ" if row['dekont_no'] and str(row['dekont_no']).strip() else "❌ BEKLİYOR"
                        
                        # Update Keşif Record
                        cursor.execute("""
                            UPDATE kesif_kayitlari SET dekont_no=?, payment_date=?, bank_info=?, status=? WHERE id=?
                        """, (row['dekont_no'], row['payment_date'], row['bank_info'], new_status, row['id']))
                        
                        # Sync with Protocol Table (by linked project name)
                        cursor.execute("""
                            UPDATE protocols 
                            SET receipt_no=?, payment_date=?, bank_info=? 
                            WHERE linked_project_name=?
                        """, (row['dekont_no'], row['payment_date'], row['bank_info'], proj_name))
                        
                conn_sett.commit()
                st.rerun()
            
            # 2. Handle Deletion
            if st.button("🗑️ Seçili Kaydı Sil", type="primary"):
                selected_ids = edited_saved[edited_saved["Seç"] == True]["id"].tolist()
                if selected_ids:
                    cursor = conn_sett.cursor()
                    placeholders = ",".join(["?"] * len(selected_ids))
                    cursor.execute(f"DELETE FROM kesif_kayitlari WHERE id IN ({placeholders})", selected_ids)
                    conn_sett.commit()
                    st.success(f"{len(selected_ids)} kayıt silindi.")
                    st.rerun()
                else:
                    st.warning("Silmek için önce listeden seçim yapmalısınız.")
        else:
            st.info("Henüz kayıtlı bir keşif bulunmuyor.")
 
    # Selection: Project or Kayıt No
    mode = st.radio("Seçim Yöntemi", ["Proje Bazlı", "Kayıt No Bazlı"], horizontal=True)
    
    sel_proj = ""
    if mode == "Proje Bazlı":
        projects = pd.read_sql("SELECT DISTINCT proje FROM samples WHERE proje IS NOT NULL AND proje != ''", conn_data)
        sel_proj = st.selectbox("Proje Seçin / Girin", [""] + projects['proje'].tolist())
        query = f"SELECT * FROM samples WHERE proje='{sel_proj}'"
    else:
        knos = pd.read_sql("SELECT DISTINCT kayit_no FROM samples ORDER BY id DESC LIMIT 500", conn_data)
        sel_kno = st.selectbox("Kayıt No Seçin", knos['kayit_no'].tolist())
        query = f"SELECT * FROM samples WHERE kayit_no='{sel_kno}'"
        # If it's Kayıt No based, we use the Kayıt No as the identifier for settings/extras if needed,
        # or we find the project name associated with it.
        sel_proj = sel_kno # Fallback to Kayıt No as the 'project identifier' for this view
        
    df_samples = pd.read_sql(query, conn_data)
    
    is_admin = st.session_state.get('user_role') == 'admin'
 
    # 📈 Birim Fiyat Yönetimi Section (Admins Only)
    if is_admin:
        with st.expander("📈 Birim Fiyat Yönetimi (Yıllık Tablo Güncelleme)"):
            st.info("Bu bölümden yeni yılın birim fiyatlarını Excel veya PDF dosyasından yükleyebilir veya mevcut fiyatları elle düzenleyebilirsiniz.")
            
            up_col1, up_col2 = st.columns([1, 2])
            with up_col1:
                target_year = st.number_input("Güncellenecek Yıl", value=2026, step=1)
                import_method = st.radio("Yükleme Yöntemi", ["Dosya Yükle (Excel/PDF)", "Metin Yapıştır (PDF'ten Kopyala)"])
                
                uploaded_file = None
                pasted_text = None
                
                if "Dosya" in import_method:
                    uploaded_file = st.file_uploader("Dosya Seç (Excel veya PDF)", type=["xlsx", "pdf"])
                else:
                    pasted_text = st.text_area("PDF'ten kopyaladığınız metni buraya yapıştırın", height=200, 
                                              help="PDF'teki tüm metni seçip (Ctrl+A) buraya yapıştırabilirsiniz.")
            
            # Show existing prices for the selected year
            df_prices = pd.read_sql(f"SELECT * FROM kesif_birim_fiyatlar WHERE year={target_year}", conn_sett)
            
            if uploaded_file or pasted_text:
                try:
                    df_up = pd.DataFrame()
                    if uploaded_file:
                        if uploaded_file.name.endswith(".pdf"):
                            with pdfplumber.open(uploaded_file) as pdf:
                                all_rows = []
                                text_rows = []
                                # --- Strategy 1: Table Extraction ---
                                for page in pdf.pages:
                                    table = page.extract_table()
                                    if table:
                                        # Normalize row lengths and filter empty
                                        table = [r for r in table if r and any(cell is not None and str(cell).strip() != "" for cell in r)]
                                        all_rows.extend(table)

                                    # Table output can look valid while columns are shifted; always scan text too.
                                    text_rows.extend(extract_price_rows_from_text(page.extract_text() or ""))

                                # Prefer validated text rows for PDFs, then fall back to the extracted table.
                                if text_rows:
                                    deduplicated_rows = {}
                                    for row in text_rows:
                                        deduplicated_rows[normalize_poz_no(row[0])] = row
                                    df_up = pd.DataFrame(
                                        list(deduplicated_rows.values()),
                                        columns=["Poz No", "Tanım", "Birim", "Birim Fiyat"]
                                    )
                                    st.success(f"PDF içinden {len(df_up)} geçerli poz-fiyat satırı bulundu.")
                                elif not all_rows or (all_rows and len(all_rows[0]) < 3):
                                    st.error("PDF içinde poz numarası ve birim fiyat birlikte okunamadı.")
                                else:
                                    # Normal Table Display
                                    max_cols = max(len(r) for r in all_rows)
                                    normalized_rows = []
                                    for r in all_rows:
                                        padding = [None] * (max_cols - len(r))
                                        normalized_rows.append(r + padding)
                                    # Ensure unique headers to avoid Pandas error
                                    raw_headers = normalized_rows[0]
                                    unique_headers = []
                                    seen_names = {}
                                    for i, h in enumerate(raw_headers):
                                        name = str(h).strip() if h is not None and str(h).strip() != "" else f"Sutun_{i+1}"
                                        if name in seen_names:
                                            seen_names[name] += 1
                                            name = f"{name}_{seen_names[name]}"
                                        else:
                                            seen_names[name] = 0
                                        unique_headers.append(name)
                                    
                                    df_up = pd.DataFrame(normalized_rows[1:], columns=unique_headers)
                        else:
                            df_up = pd.read_excel(uploaded_file)
                    
                    elif pasted_text:
                        st.info("Yapıştırılan metin işleniyor...")
                        text_rows = extract_price_rows_from_text(pasted_text)
                        if text_rows:
                            df_up = pd.DataFrame(text_rows, columns=["Poz No", "Tanım", "Birim", "Birim Fiyat"])
                        else:
                            st.error("Yapıştırılan metinde Poz numarası formatında veri bulunamadı.")
                    
                    if not df_up.empty:
                        st.write("Dosya İçeriği (Önizleme):")
                        st.dataframe(df_up.head(10), use_container_width=True)
                    
                        st.warning(f"⚠️ Bu işlem {target_year} yılına ait varsa tüm eski fiyatları silecek ve yenilerini ekleyecektir.")
                        
                        # 4 Explicit Selectboxes for Mapping
                        cols = df_up.columns.tolist()
                        
                        # Attempt to auto-detect
                        def guess_col(keywords, cols):
                            for i, c in enumerate(cols):
                                if any(k in str(c).upper() for k in keywords):
                                    return i
                            return 0

                        st.markdown("#### 🛠️ Sütun Eşleştirmesi")
                        st.info("Lütfen PDF/Excel dosyasındaki sütunların hangisinin ne olduğunu seçin.")
                        
                        mc1, mc2, mc3, mc4 = st.columns(4)
                        mapped_poz = mc1.selectbox("Poz No Sütunu", cols, index=guess_col(["POZ", "NO"], cols))
                        mapped_desc = mc2.selectbox("Tanım Sütunu", cols, index=guess_col(["TANIM", "ADI", "AÇIKLAMA"], cols))
                        mapped_unit = mc3.selectbox("Birim Sütunu", cols, index=guess_col(["BİRİM"], cols))
                        mapped_price = mc4.selectbox("Birim Fiyat Sütunu", cols, index=guess_col(["FİYAT", "TUTAR", "BEDEL"], cols))

                        def clean_price(val):
                            return clean_price_value(val)

                        if st.button(f"🔴 {target_year} YILI VERİLERİNİ SİSTEME AKTAR"):
                            valid_rows = {}
                            for _, row in df_up.iterrows():
                                p_no = normalize_poz_no(row[mapped_poz])
                                # DEEP SCAN: If Price column is empty, look for it in other columns
                                price_raw = row[mapped_price]
                                price = clean_price(price_raw)
                                
                                if price == 0:
                                    # Scan all columns for this row
                                    full_row_text = " ".join([str(v) for v in row.values if v is not None])
                                    deep_matches = re.findall(r'(\d+[\.,]?\d*[\.,]\d{1,2})', full_row_text)
                                    if deep_matches:
                                        price = clean_price(deep_matches[-1])
                                
                                desc = str(row[mapped_desc]).strip()
                                unit = str(row[mapped_unit]).strip() if mapped_unit else "Adet"
                                
                                if not is_valid_poz_no(p_no) or price <= 0:
                                    continue
                                valid_rows[p_no] = (p_no, desc or "-", unit or "Adet", price)

                            if not valid_rows:
                                st.error("Geçerli poz numarası ve birim fiyat bulunamadı. Mevcut yıl fiyatları korunmuştur.")
                                conn_data.close()
                                conn_sett.close()
                                return

                            cursor = conn_sett.cursor()
                            cursor.execute("DELETE FROM kesif_birim_fiyatlar WHERE year=?", (int(target_year),))
                            for p_no, desc, unit, price in valid_rows.values():
                                cursor.execute("""
                                    INSERT INTO kesif_birim_fiyatlar (poz_no, description, unit, price, year, lab_type)
                                    VALUES (?, ?, ?, ?, ?, ?)
                                """, (p_no, desc, unit, price, target_year, "GENEL"))
                            
                            conn_sett.commit()
                            st.success(f"✅ {target_year} yılı için {len(valid_rows)} adet birim fiyat başarıyla yüklendi!")
                            st.rerun()
                except Exception as e:
                    st.error(f"Hata: {str(e)}")
            
            st.markdown("---")
            st.markdown(f"#### 📋 {target_year} Yılı Mevcut Fiyat Listesi")
            if not df_prices.empty:
                edited_prices = st.data_editor(df_prices, key=f"price_editor_{target_year}", use_container_width=True, num_rows="dynamic")
                if st.button(f"💾 {target_year} TABLOSUNU KAYDET", key=f"save_prices_{target_year}"):
                    cursor = conn_sett.cursor()
                    cursor.execute(f"DELETE FROM kesif_birim_fiyatlar WHERE year={target_year}")
                    for _, row in edited_prices.iterrows():
                         cursor.execute("""
                            INSERT INTO kesif_birim_fiyatlar (poz_no, description, unit, price, year, lab_type)
                            VALUES (?, ?, ?, ?, ?, ?)
                        """, (row['poz_no'], row['description'], row['unit'], row['price'], target_year, row['lab_type']))
                    conn_sett.commit()
                    st.success("Fiyat tablosu güncellendi.")
                    st.rerun()
            else:
                st.warning(f"Bu yıl ({target_year}) için henüz kayıt bulunmuyor.")

    if sel_proj == "":
        st.info("Lütfen bir seçim yapın.")
        conn_data.close()
        conn_sett.close()
        return

    if df_samples.empty:
        st.warning("Seçilen kriterlere uygun kayıt bulunamadı.")
        conn_data.close()
        conn_sett.close()
        return

    # Load project settings if existing from Settings DB
    proj_settings = pd.read_sql(f"SELECT * FROM kesif_ayarlari WHERE project_name='{sel_proj}'", conn_sett)
    
    # 1. Load settings from DB into variables first (Calculations must work for everyone)
    if not proj_settings.empty:
        s = proj_settings.iloc[0]
        year = int(s['year'])
        protocol_rate = float(s['protocol_rate'])
        acc_rate = float(s['acc_rate'])
        project_coeff_rate = float(s['project_coeff_rate']) if 'project_coeff_rate' in proj_settings.columns else 0.0
        vat_rate = float(s['vat_rate'])
        stamp_tax_rate = float(s['stamp_tax_rate']) if 'stamp_tax_rate' in proj_settings.columns else 0.00948
        turkak_rate = float(s['turkak_rate']) if 'turkak_rate' in proj_settings.columns else 1.0
    else:
        year, protocol_rate, acc_rate, project_coeff_rate, vat_rate, stamp_tax_rate, turkak_rate = 2025, 10.0, 10.0, 0.0, 20.0, 0.00948, 1.0


    # 2. Show Configuration UI only to Admins
    if is_admin:
        with st.expander("⚙️ Keşif Ayarları", expanded=False):
            c_y, c_p, c_a, c_pc, c_v, c_s, c_t = st.columns(7)
            year = c_y.selectbox("Birim Fiyat Yılı", [2025, 2024], index=0 if year == 2025 else 1)
            protocol_rate = c_p.number_input("Protokol (%)", value=protocol_rate, step=0.5)
            acc_rate = c_a.number_input("Akreditasyon (%)", value=acc_rate, step=0.5)
            project_coeff_rate = c_pc.number_input("Proje Katsayısı (%)", value=project_coeff_rate, step=0.5)
            vat_rate = c_v.number_input("KDV (%)", value=vat_rate, step=0.5)
            stamp_tax_rate_permil = c_s.number_input("Damga Ver. (‰)", value=float(stamp_tax_rate*1000), step=0.01)
            stamp_tax_rate = stamp_tax_rate_permil / 1000.0
            turkak_rate = c_t.number_input("Türkak Payı (%)", value=turkak_rate, step=0.1)
        
    # Load or set default formulas once
    def_c = "=(P21*(1+{acc_rate}/100)+P22)*(1+{protocol_rate}/100)*(1+{project_coeff_rate}/100)"
    def_d = "=ROUND(P23*{vat_rate}/100, 2)"
    def_e = "=ROUND(P23*{stamp_tax_rate}, 2)"
    def_f = "=ROUND(P21*{turkak_rate}/100, 2)"
    def_g = "=P23+P24+P25+P26"

    formula_c = proj_settings.iloc[0]['formula_c'] if not proj_settings.empty and proj_settings.iloc[0]['formula_c'] else def_c
    formula_d = proj_settings.iloc[0]['formula_d'] if not proj_settings.empty and proj_settings.iloc[0]['formula_d'] else def_d
    formula_e = proj_settings.iloc[0]['formula_e'] if not proj_settings.empty and proj_settings.iloc[0]['formula_e'] else def_e
    formula_f = proj_settings.iloc[0]['formula_f'] if not proj_settings.empty and proj_settings.iloc[0]['formula_f'] else def_f
    formula_g = proj_settings.iloc[0]['formula_g'] if not proj_settings.empty and proj_settings.iloc[0]['formula_g'] else def_g

    # Admin Formula Settings
    if is_admin:
        with st.expander("🔐 Yönetici Formül Ayarları"):
            admin_pass = st.text_input("Yönetici Şifresi", type="password")
            if admin_pass == "admin123": # For demo purposes
                st.info("Formüllerde P21(A), P22(B), P23(C)... hücrelerini ve {protocol_rate} gibi değişkenleri kullanabilirsiniz.")
                
                formula_c = st.text_input("[C] Toplam Protokollü Bedel Formülü", value=formula_c)
                formula_d = st.text_input("[D] KDV Formülü", value=formula_d)
                formula_e = st.text_input("[E] Damga Vergisi Formülü", value=formula_e)
                formula_f = st.text_input("[F] Türkak Payı Formülü", value=formula_f)
                formula_g = st.text_input("[G] Genel Toplam Formülü", value=formula_g)
                
                if st.button("🔄 Formülleri Varsayılana Döndür", key="reset_formulas"):
                    cursor = conn_sett.cursor()
                    cursor.execute("""
                        UPDATE kesif_ayarlari SET formula_c=?, formula_d=?, formula_e=?, formula_f=?, formula_g=?
                        WHERE project_name=?
                    """, (def_c, def_d, def_e, def_f, def_g, sel_proj))
                    conn_sett.commit()
                    st.info("Formüller varsayılan değerlere çekildi ve kaydedildi.")
                    st.rerun()
                    
                # Transparency: Show how placeholders are replaced
                debug_f = formula_c
                for k,v in {"{acc_rate}":acc_rate, "{protocol_rate}":protocol_rate, "{project_coeff_rate}":project_coeff_rate}.items():
                    debug_f = debug_f.replace(k, f"({v})")
                st.caption(f"Denetim Formülü: {debug_f}")
            else:
                if admin_pass: st.error("Hatalı Şifre")

    if is_admin:
        if st.button("Ayarları Kaydet", key="save_settings_btn"):
            cursor = conn_sett.cursor()
            cursor.execute("""
                INSERT INTO kesif_ayarlari (project_name, year, protocol_rate, acc_rate, vat_rate, stamp_tax_rate, project_coeff_rate, turkak_rate, 
                                            formula_c, formula_d, formula_e, formula_f, formula_g)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(project_name) DO UPDATE SET
                year=excluded.year, protocol_rate=excluded.protocol_rate, 
                acc_rate=excluded.acc_rate, vat_rate=excluded.vat_rate,
                stamp_tax_rate=excluded.stamp_tax_rate,
                project_coeff_rate=excluded.project_coeff_rate,
                turkak_rate=excluded.turkak_rate,
                formula_c=excluded.formula_c, formula_d=excluded.formula_d, formula_e=excluded.formula_e,
                formula_f=excluded.formula_f, formula_g=excluded.formula_g
            """, (sel_proj, year, protocol_rate, acc_rate, vat_rate, stamp_tax_rate, project_coeff_rate, turkak_rate, 
                  formula_c, formula_d, formula_e, formula_f, formula_g))
            conn_sett.commit()
            st.success("Ayarlar kaydedildi.")

    # Extraction Logic
    experiments_found = df_samples['deney_adi'].tolist()
    
    from collections import Counter
    exp_counts = Counter(experiments_found)
    
    # Prices and Mapping from Settings DB
    df_prices = pd.read_sql(f"SELECT * FROM kesif_birim_fiyatlar WHERE year={year}", conn_sett)
    df_test_defs = pd.read_sql("SELECT test_name, poz_no, is_accredited FROM test_durations WHERE poz_no IS NOT NULL AND poz_no != ''", conn_sett)
    
    # Load project-specific accreditation overrides from Settings DB
    df_acc_overrides = pd.read_sql(f"SELECT poz_no, is_accredited FROM kesif_akredite_durumlari WHERE project_name='{sel_proj}'", conn_sett)
    acc_overrides = {r['poz_no'].strip(): bool(r['is_accredited']) for _, r in df_acc_overrides.iterrows()}
    
    # Build robust mapping
    df_test_defs['norm_name'] = df_test_defs['test_name'].str.strip().str.lower()
    robust_mapping = {}
    for _, r in df_test_defs.iterrows():
        # Use override if exists, otherwise default
        p_code = r['poz_no'].strip()
        is_acc = acc_overrides.get(p_code, bool(r['is_accredited']))
        robust_mapping[r['norm_name']] = (p_code, is_acc, r['test_name'])
    
    kesif_items = []
    # Normalize sample experiments for matching
    df_samples['norm_name'] = df_samples['deney_adi'].str.strip().str.lower()
    exp_counts = Counter(df_samples['norm_name'].tolist())

    for norm_name, count in exp_counts.items():
        mapping_data = robust_mapping.get(norm_name)
        if mapping_data:
            poz, is_acc, orig_test_name = mapping_data
            # Flexible price lookup
            price_row = df_prices[df_prices['poz_no'].apply(normalize_poz_no) == normalize_poz_no(poz)]
            
            p_price = 0.0
            p_desc = f"Fiyat Bulunamadı ({poz})"
            p_unit = "-"
            p_code = poz
            
            if not price_row.empty:
                item = price_row.iloc[0]
                p_price = float(item['price'])
                p_desc = item['description']
                p_unit = item['unit']
                p_code = item['poz_no']
            
            kesif_items.append({
                "Poz No": p_code,
                "Tanım": p_desc,
                "Birim": p_unit,
                "Miktar": float(count),
                "Birim Fiyat": p_price,
                "Tutar": float(count * p_price),
                "is_accredited": is_acc
            })
    
    # Session state for extra items
    if 'kesif_extra_items' not in st.session_state:
        st.session_state.kesif_extra_items = {}
    
    # Handle project change
    if 'last_sel_proj' not in st.session_state or st.session_state.last_sel_proj != sel_proj:
        st.session_state.last_sel_proj = sel_proj
        # Load extra items from Settings DB
        db_extras = pd.read_sql(f"SELECT * FROM kesif_ek_kalemler WHERE project_name='{sel_proj}'", conn_sett)
        st.session_state.kesif_extra_items[sel_proj] = []
        for _, r in db_extras.iterrows():
            st.session_state.kesif_extra_items[sel_proj].append({
                "Poz No": r['poz_no'], "Tanım": r['description'], "Birim": r['unit'],
                "Miktar": r['quantity'], "Birim Fiyat": r['price'], "Tutar": r['quantity'] * r['price'],
                "is_accredited": bool(r.get('is_accredited', 0)),
                "Original_Test_Name": orig_test_name if 'orig_test_name' in locals() else ""
            })

    # UI for Extra Items
    with st.expander("➕ Keşfe Manuel Kalem Ekle"):
        col_e1, col_e2, col_e3, col_e4 = st.columns([3, 1, 1, 1])
        with col_e1:
            all_poz_list = (df_prices['poz_no'] + " - " + df_prices['description']).tolist()
            sel_extra = st.selectbox("Birim Fiyat Listesinden Seç", ["Seçiniz..."] + all_poz_list)
        with col_e2:
            extra_qty = st.number_input("Miktar", min_value=0.1, value=1.0, step=0.1)
        with col_e3:
            is_acc_extra = st.checkbox("Akredite Mi?", value=False)
        with col_e4:
            if st.button("EKLE") and sel_extra != "Seçiniz...":
                p_code = sel_extra.split(" - ")[0]
                ex_row = df_prices[df_prices['poz_no'] == p_code].iloc[0]
                st.session_state.kesif_extra_items[sel_proj].append({
                    "Poz No": ex_row['poz_no'], "Tanım": ex_row['description'], "Birim": ex_row['unit'],
                    "Miktar": extra_qty, "Birim Fiyat": ex_row['price'], "Tutar": extra_qty * ex_row['price'],
                    "is_accredited": is_acc_extra
                })
                # Save to Settings DB immediately
                cursor = conn_sett.cursor()
                cursor.execute("INSERT INTO kesif_ek_kalemler (project_name, poz_no, description, unit, price, quantity, is_accredited) VALUES (?,?,?,?,?,?,?)",
                             (sel_proj, ex_row['poz_no'], ex_row['description'], ex_row['unit'], ex_row['price'], extra_qty, int(is_acc_extra)))
                conn_sett.commit()
                st.rerun()

    current_extras = st.session_state.kesif_extra_items.get(sel_proj, [])
    final_list = kesif_items + current_extras
    
    if final_list:
        df_final = pd.DataFrame(final_list)
        st.subheader("📊 Keşif Kalemleri")
        # Use data_editor to allow toggling is_accredited directly in the UI
        edited_df = st.data_editor(df_final, use_container_width=True, hide_index=False,
                                  column_config={"is_accredited": st.column_config.CheckboxColumn("Akredite Mi?")})
        
        # Update final_list from edited_df for the following math
        final_list = edited_df.to_dict('records')
        
        # PERSIST "Akredite Mi?" changes immediately to the Settings database
        if not edited_df.equals(df_final):
            cursor = conn_sett.cursor()
            for _, row in edited_df.iterrows():
                cursor.execute("""
                    INSERT INTO kesif_akredite_durumlari (project_name, poz_no, is_accredited)
                    VALUES (?, ?, ?)
                    ON CONFLICT(project_name, poz_no) DO UPDATE SET is_accredited=excluded.is_accredited
                """, (sel_proj, row['Poz No'], int(row['is_accredited'])))
            conn_sett.commit()
            st.rerun()
        
        # Separate Items
        acc_items = [i for i in final_list if i.get('is_accredited', False)]
        normal_items = [i for i in final_list if not i.get('is_accredited', False)]
        
        # [A] TS EN ISO/IEC 17025 AKREDİTE HİZMETLER TUTARI
        sum_a = round(sum(i['Tutar'] for i in acc_items), 2)
        
        # [B] DİĞER HİZMETLER TUTARI
        sum_b = round(sum(i['Tutar'] for i in normal_items), 2)
        
        # UI Formula Evaluation Context
        context = {
            "P21": sum_a,
            "P22": sum_b,
            "{acc_rate}": acc_rate,
            "{protocol_rate}": protocol_rate,
            "{project_coeff_rate}": project_coeff_rate,
            "{vat_rate}": vat_rate,
            "{stamp_tax_rate}": stamp_tax_rate,
            "{turkak_rate}": turkak_rate
        }
        
        # Evaluate Steps sequentially to allow P23 to reference C result, etc.
        sum_c = parse_formula_for_ui(formula_c, context)
        context["P23"] = sum_c # C
        
        vat_amt = parse_formula_for_ui(formula_d, context)
        context["P24"] = vat_amt # D
        
        stamp_amt = parse_formula_for_ui(formula_e, context)
        context["P25"] = stamp_amt # E
        
        turkak_payi = parse_formula_for_ui(formula_f, context)
        context["P26"] = turkak_payi # F
        
        grand_total = parse_formula_for_ui(formula_g, context)
        
        c1, c2 = st.columns(2)
        with c2:
            st.write(f"**[A] TS EN ISO/IEC 17025 AKREDİTE HİZMETLER TUTARI:** {sum_a:,.2f} ₺")
            st.write(f"**[B] DİĞER HİZMETLER TUTARI:** {sum_b:,.2f} ₺")
            
            st.write(f"**[C] TOPLAM PROTOKOLLÜ HİZMET BEDELİ:** {sum_c:,.2f} ₺")
            st.write(f"**[D] KDV:** {vat_amt:,.2f} ₺")
            st.write(f"**[E] DAMGA VERGİSİ:** {stamp_amt:,.2f} ₺")
            
            if acc_items and turkak_rate > 0:
                st.write(f"**[F] TÜRKAK PAYI:** {turkak_payi:,.2f} ₺")
            
            st.markdown("---")
            if st.button("📌 Keşfi Takibe Al / Kaydet", use_container_width=True):
                cursor = conn_sett.cursor()
                now = datetime.now().strftime("%d/%m/%Y %H:%M")
                cursor.execute("""
                    INSERT INTO kesif_kayitlari (project_name, year, total_amount, created_at)
                    VALUES (?, ?, ?, ?)
                """, (sel_proj, year, grand_total, now))
                conn_sett.commit()
                st.success("Keşif başarıyla takibe alındı.")
                st.rerun()
                
            st.markdown("---")
            st.markdown(f"### **GENEL TOPLAM:** {grand_total:,.2f} ₺")
            
            # Show formulas for transparency
            with st.expander("🔍 Hesaplama Formülleri"):
                st.code(f"C: {formula_c}\nD: {formula_d}\nE: {formula_e}\nF: {formula_f}\nG: {formula_g}")

        if st.button("🗑️ Manuel Kalemleri Temizle"):
             cursor = conn_sett.cursor()
             cursor.execute("DELETE FROM kesif_ek_kalemler WHERE project_name=?", (sel_proj,))
             conn_sett.commit()
             st.session_state.kesif_extra_items[sel_proj] = []
             st.rerun()

        st.markdown("---")
        st.subheader("✍️ Rapor İmza Bilgileri")
        sig_col1, sig_col2, sig_col3 = st.columns(3)
        with sig_col1:
            sig_d = st.text_input("Düzenleyen İsmi", value=st.session_state.get('username', ''))
            sig_d_uv = st.text_input("Düzenleyen Ünvanı", value="", key="sig_d_uv")
        with sig_col2:
            sig_k = st.text_input("Kontrol Eden İsmi", value="")
            sig_k_uv = st.text_input("Kontrol Eden Ünvanı", value="", key="sig_k_uv")
        with sig_col3:
            sig_o = st.text_input("Onaylayan İsmi", value="")
            sig_o_uv = st.text_input("Onaylayan Ünvanı", value="", key="sig_o_uv")

        # --- Word Export Section (Only show after Excel is generated) ---
        if st.session_state.get(f'kesif_excel_generated_{sel_proj}', False):
            st.markdown("---")
            st.subheader("📄 Word Protokol Oluşturma")
        
            with st.expander("🛠️ Word Şablonu ve Ayarları", expanded=True):
                template_path = "PROTOKOL.docx"
                if not os.path.exists(template_path):
                    st.error(f"⚠️ '{template_path}' dosyası ana klasörde bulunamadı!")
                    st.info("Lütfen 'PROTOKOL.docx' adındaki şablon dosyasını uygulamanın çalıştığı klasöre yükleyin.")
                else:
                    st.success(f"✅ Şablon bulundu: {template_path}")
                    
                    # Template Inspection
                    if st.button("🔍 Şablon İçeriğini ve Yer Tutucuları Tara"):
                        try:
                            import zipfile
                            found_placeholders = []
                            with zipfile.ZipFile(template_path) as z:
                                xml_content = z.read('word/document.xml').decode('utf-8')
                                # Simple regex to find {Word} style placeholders
                                found_placeholders = re.findall(r'\{[^{}]+\}', xml_content)
                                
                                st.write("**Şablon İçeriği (Özet):**")
                                clean_text = re.sub('<[^>]+>', '', xml_content)
                                st.caption(clean_text[:500] + "...")
                                
                                if found_placeholders:
                                    st.write("**Bulunan Yer Tutucular:**", list(set(found_placeholders)))
                                else:
                                    st.warning("Şablonda {Etiket} formatında yer tutucu bulunamadı. Lütfen Word dosyasını düzenleyip ilgili yerlere {Firma}, {Tutar} vb. etiketler ekleyin.")
                        except Exception as e:
                            st.error(f"Şablon okuma hatası: {e}")

                    # Input Fields for Export
                    st.markdown("#### 📝 Protokol Bilgileri")
                    w_col1, w_col2, w_col3 = st.columns(3)
                    w_firm = w_col1.text_input("Firma Adı", value=df_samples['firma'].iloc[0] if not df_samples.empty else "")
                    w_tax = w_col2.text_input("Vergi No / TC")
                    w_date = w_col3.date_input("Protokol Tarihi", value=datetime.now())
                    if st.button("📄 WORD DOSYASI OLUŞTUR (İndir)", type="primary"):
                        try:
                            # Prepare Data - Directly use already computed UI totals to avoid desync/errors
                            costs = {
                                "base_cost": sum_c,
                                "kdv_amount": vat_amt,
                                "stamp_tax": stamp_amt,
                                "turkak_fee": turkak_payi,
                                "total_amount": grand_total
                            }
                            data_map = {
                                "{Firma}": w_firm,
                                "FİRMA ADI GELECEK": w_firm, # Support user's existing text
                                "{VergiNo}": w_tax,
                                "{Vergi}": w_tax,
                                "{TC}": w_tax,
                                "{Tc}": w_tax,
                                "{VKN}": w_tax,
                                "{Tarih}": w_date.strftime("%d.%m.%Y"),
                                "{Proje}": sel_proj,
                                "{HizmetBedeli}": f"{costs['base_cost']:,.2f}",
                                "{KDV}": f"{costs['kdv_amount']:,.2f}",
                                "{Damga}": f"{costs['stamp_tax']:,.2f}",
                                "{Turkak}": f"{costs['turkak_fee']:,.2f}",
                                "{Toplam}": f"{costs['total_amount']:,.2f}",
                                # Add generic letters matches
                                "{C}": f"{costs['base_cost']:,.2f}",
                                "{D}": f"{costs['kdv_amount']:,.2f}",
                                "{E}": f"{costs['stamp_tax']:,.2f}",
                                "{F}": f"{costs['turkak_fee']:,.2f}",
                            }
                            
                            # Generate
                            import zipfile
                            
                            # Read template
                            with open(template_path, 'rb') as f:
                                template_bytes = f.read()
                                
                            # Load into ZipFile in memory to modify
                            bio_in = io.BytesIO(template_bytes)
                            bio_out = io.BytesIO()
                            
                            with zipfile.ZipFile(bio_in, 'r') as zin:
                                with zipfile.ZipFile(bio_out, 'w') as zout:
                                    zout.comment = zin.comment # preserve comment
                                    for item in zin.infolist():
                                        if item.filename == 'word/document.xml':
                                            xml_content = zin.read(item.filename).decode('utf-8')
                                            
                                            # 1. Clean up split placeholders (e.g. {<w:t>Vergi</w:t>No})
                                            # This regex finds { followed by any number of XML tags and content, ending with }
                                            # And collapses it if it matches a known key pattern
                                            
                                            # Simple brute-force for specific known troublesome keys
                                            # Remove XML tags inside curly braces if they break a keyword
                                            # Example: {<w..>Vergi<w..>No<w..>} -> {VergiNo}
                                            
                                            def clean_xml_tags_in_placeholders(match):
                                                # Remove all <...> tags from within the match
                                                clean = re.sub(r'<[^>]+>', '', match.group(0))
                                                return clean

                                            # Regex to find curly braces with potential tags inside
                                            # \{[^{}]*\} might match {<w:t>Vergi</w:t>}
                                            xml_content = re.sub(r'\{[^{}]+\}', clean_xml_tags_in_placeholders, xml_content)
                                            
                                            # 2. Now perform standard replacement
                                            for key, val in data_map.items():
                                                xml_content = xml_content.replace(key, str(val))
                                                
                                            zout.writestr(item, xml_content.encode('utf-8'))
                                        else:
                                            zout.writestr(item, zin.read(item.filename))
                                            
                            bio_out.seek(0)
                            
                            # Save to database
                            file_bytes_for_db = bio_out.getvalue()
                            save_word_protokol(
                                sel_proj, w_firm, w_tax, 
                                w_date.strftime("%Y-%m-%d"), 
                                costs, 
                                file_bytes_for_db
                            )
                            
                            st.download_button(
                                label="⬇️ Oluşturulan Word Dosyasını İndir",
                                data=file_bytes_for_db,
                                file_name=f"Protokol_{sel_proj}.docx",
                                mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document"
                            )
                            st.success("✅ Word dosyası oluşturuldu ve kaydedildi! İndirebilirsiniz.")
                            
                        except Exception as e:
                            st.error(f"Hata oluştu: {e}")
        else:
            st.info("ℹ️ Word Protokol oluşturmak için önce Excel Keşif Raporunu oluşturun.")
        
        # --- NEW: Quick Add to Protocol List ---
        if st.session_state.get(f'kesif_excel_generated_{sel_proj}', False):
            st.markdown("---")
            st.subheader("🔗 Protokol Listesine Hızlı Ekleme")
            st.info("Bu bölümdeki butonu kullanarak, yukarıdaki hesaplanmış verileri ve firma bilgilerini doğrudan 'Protokol Listesi'ne aktarabilirsiniz. Tekrar veri girmenize gerek kalmaz.")
            
            if st.button("📋 Verileri Protokol Listesine Kaydet", type="secondary", use_container_width=True):
                try:
                    # 1. Gather Data - Use already computed UI totals
                    costs = {
                        "base_cost": sum_c,
                        "kdv_amount": vat_amt,
                        "stamp_tax": stamp_amt,
                        "turkak_fee": turkak_payi,
                        "total_amount": grand_total
                    }
                    # Use the inputs from the Word section if available, else fallback
                    final_firm = w_firm if 'w_firm' in locals() else df_samples['firma'].iloc[0]
                    final_date = w_date.strftime("%Y-%m-%d") if 'w_date' in locals() else datetime.now().strftime("%Y-%m-%d")
                    job_desc = f"{sel_proj} - {final_firm} - Laboratuvar Deney Ücreti"
                    
                    # 2. Insert into Settings DB
                    cursor = conn_sett.cursor()
                    # Use YYYY-MM for the month column to match the new schema
                    current_period = datetime.now().strftime("%Y-%m")
                    
                    cursor.execute("""
                        INSERT INTO protocols (
                            firm, job_description, protocol_date, 
                            base_cost, kdv_amount, stamp_tax, turkak_fee, total_amount,
                            payment_date, month, linked_project_name, is_archived
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0)
                    """, (
                        final_firm, 
                        job_desc, 
                        final_date,
                        float(costs.get('base_cost', 0.0)), 
                        float(costs.get('kdv_amount', 0.0)), 
                        float(costs.get('stamp_tax', 0.0)), 
                        float(costs.get('turkak_fee', 0.0)), 
                        float(costs.get('total_amount', 0.0)),
                        final_date, # Default payment date to protocol date so it's not empty
                        current_period, 
                        sel_proj
                    ))
                    conn_sett.commit()
                    st.success(f"✅ '{sel_proj}' projesi Protokol Listesine başarıyla eklendi! 'Protokol ve Maliyet Takip' sekmesinden görüntüleyebilirsiniz.")
                except Exception as e:
                    import traceback
                    print(f"ERROR during transfer: {traceback.format_exc()}")
                    st.error(f"Kayıt sırasında hata oluştu: {e}")


        st.markdown("---")
        
        if st.button("📥 Excel Raporu Oluştur", use_container_width=True):
            # Extract metadata
            knos = df_samples['kayit_no'].unique().tolist()
            knos_str = ", ".join(map(str, knos))
            servis = df_samples['ilgili_servis'].iloc[0] if 'ilgili_servis' in df_samples.columns and not df_samples['ilgili_servis'].empty else ""
            cinsi = df_samples['cins'].iloc[0] if 'cins' in df_samples.columns and not df_samples['cins'].empty else ""
            
            # Template path logic - Check local first, then default to older logic or warn
            possible_paths = [
                "Kesif.xlsx", # Current directory
                os.path.join(os.getcwd(), "Kesif.xlsx"),
                r'C:\Users\hsynd\Downloads\Kesif.xlsx' # Fallback
            ]
            
            template_path = None
            for p in possible_paths:
                if os.path.exists(p):
                    template_path = p
                    break
            
            if template_path:
                output = export_to_excel(
                    final_list, sel_proj, year, protocol_rate, acc_rate, vat_rate, 
                    template_path, knos_str, servis, stamp_tax_rate, project_coeff_rate, 
                    cinsi, turkak_rate=turkak_rate,
                    f_c=formula_c, f_d=formula_d, f_e=formula_e, f_f=formula_f, f_g=formula_g,
                    sig_d=sig_d, sig_k=sig_k, sig_o=sig_o,
                    sig_d_uv=sig_d_uv, sig_k_uv=sig_k_uv, sig_o_uv=sig_o_uv
                )
                st.download_button(
                    label="💾 Excel Dosyasını İndir",
                    data=output,
                    file_name=f"Kesif_{sel_proj}_{year}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                # Mark excel as generated for this project
                st.session_state[f'kesif_excel_generated_{sel_proj}'] = True
            else:
                st.error(f"Taslak dosya bulunamadı: {template_path}")
    else:
        st.info("Bu proje için henüz bir deney kaydı veya manuel kalem bulunmuyor.")

    conn_data.close()
    conn_sett.close()

def export_to_excel(items, project_name, year, p_rate, a_rate, v_rate, template_path, 
                    kayit_no="", lab_type="", stamp_tax_rate=0.00948, p_coeff_rate=0.0, 
                    cinsi="", turkak_rate=1.0, f_c="", f_d="", f_e="", f_f="", f_g="",
                    sig_d="", sig_k="", sig_o="",
                    sig_d_uv="", sig_k_uv="", sig_o_uv=""):
    # Try to load the workbook with error handling
    try:
        wb = openpyxl.load_workbook(template_path)
    except Exception as e:
        raise FileNotFoundError(f"Şablon dosyası (Kesif.xlsx) bulunamadı veya açılamadı! Lütfen dosyanın proje klasöründe olduğundan emin olun. Hata: {e}")

    ws = wb.active
    
    # Separate Items
    normal_items = [i for i in items if not i.get('is_accredited', False)]
    acc_items = [i for i in items if i.get('is_accredited', False)]

    # Metadata values
    ws['E10'] = project_name
    ws['E6'] = datetime.now().strftime('%d.%m.%Y')
    ws['E4'] = kayit_no
    ws['E5'] = lab_type # Gönderen
    ws['E9'] = kayit_no # Laboratuar No
    ws['E8'] = cinsi
    
    # Combined Item Writing (Rows 13-20)
    # Track counting to separate A and D sums
    n_count = len(normal_items)
    a_count = len(acc_items)
    all_final_items = normal_items + acc_items
    
    row_idx = 13
    for i, item in enumerate(all_final_items, 1):
        if row_idx >= 21: break # Template boundary
        ws.cell(row=row_idx, column=2, value=i)
        ws.cell(row=row_idx, column=3, value=item['Poz No'])
        ws.cell(row=row_idx, column=5, value=item['Tanım'])
        ws.cell(row=row_idx, column=12, value=item['Birim'])
        ws.cell(row=row_idx, column=13, value=item['Miktar'])
        ws.cell(row=row_idx, column=14, value=item['Birim Fiyat'])
        ws.cell(row=row_idx, column=16, value=item['Tutar'])
        row_idx += 1

    # DYNAMIC RANGES BASED ON DATA
    # DYNAMIC RANGES BASED ON DATA
    # B = Sum of Normal Items
    last_normal_row = 12 + n_count
    # A = Sum of Accredited Items
    first_acc_row = 13 + n_count
    last_acc_row = 12 + n_count + a_count
    
    # PDF-STYLE COMPACT MAPPING
    # [A] TS EN ISO/IEC 17025 AKREDİTE HİZMETLER TUTARI
    # [B] DİĞER HİZMETLER TUTARI
    # [C] TOPLAM PROTOKOLLÜ HİZMET BEDELİ
    # [D] KDV
    # [E] DAMGA VERGİSİ
    # [F] TÜRKAK PAYI
    # [G] GENEL TOPLAM
    # PDF-STYLE COMPACT MAPPING (Matched to Template '10' rows 21-27)
    logical_map = {
        'A': 21, 'B': 22, 'C': 23, 'D': 24,
        'E': 25, 'F': 26, 'G': 27
    }

    # SUPER-AGGRESSIVE CLEANING (Rows 21-27 for Summary, 29-30 for Signatures)
    # We clear content but keep Row 28 (Labels) and formatting
    for r in range(21, 31):
        if r == 28: continue # SKIP LABELS
        for col in [2, 3, 7, 10, 13, 16]: # Target key columns to clear
            cell = ws.cell(row=r, column=col)
            if isinstance(cell, openpyxl.cell.cell.MergedCell):
                for m_range in ws.merged_cells.ranges:
                    if cell.coordinate in m_range:
                        ws.cell(row=m_range.min_row, column=m_range.min_col).value = None
            else:
                cell.value = None
        
        # Only hide summary calculation rows (21-26) initially, G (27) and signatures stay
        if 21 <= r <= 26:
            ws.row_dimensions[r].hidden = True

    bold_font = Font(bold=True)
    def safe_write_logic(logic_letter, formula_or_val, label=None):
        if logic_letter not in logical_map: return
        r = logical_map[logic_letter]
        if isinstance(formula_or_val, str) and formula_or_val.startswith("="):
            if f"P{r}" in formula_or_val:
                formula_or_val = formula_or_val.replace(f"P{r}", "0")
        
        v_cell = ws.cell(row=r, column=16)
        if isinstance(v_cell, openpyxl.cell.cell.MergedCell):
            for m_range in ws.merged_cells.ranges:
                if v_cell.coordinate in m_range:
                    master = ws.cell(row=m_range.min_row, column=m_range.min_col)
                    master.value = formula_or_val; master.font = bold_font
        else: v_cell.value = formula_or_val; v_cell.font = bold_font
        
        if label:
            l_cell = ws.cell(row=r, column=3)
            if isinstance(l_cell, openpyxl.cell.cell.MergedCell):
                for m_range in ws.merged_cells.ranges:
                    if l_cell.coordinate in m_range: ws.cell(row=m_range.min_row, column=m_range.min_col).value = label
            else: l_cell.value = label
        ws.row_dimensions[r].hidden = False

    # WRITE PDF-STYLE BLOCKS
    
    # [A] AKREDİTE HİZMETLER (Sum only acc rows)
    if a_count > 0:
        safe_write_logic('A', f"=SUM(P{max(13, first_acc_row)}:P{min(20, last_acc_row)})", "TS EN ISO/IEC 17025 AKREDİTE HİZMETLER TUTARI")
    else:
        safe_write_logic('A', 0, "TS EN ISO/IEC 17025 AKREDİTE HİZMETLER TUTARI")

    # [B] DİĞER HİZMETLER (Sum only normal rows)
    if n_count > 0:
        safe_write_logic('B', f"=SUM(P13:P{min(20, last_normal_row)})", "DİĞER HİZMETLER TUTARI")
    else:
        safe_write_logic('B', 0, "DİĞER HİZMETLER TUTARI")

    # WRITE CUSTOM FORMULAS
    def inject_rates(f):
        if not f: return ""
        if not f.startswith("="): f = "=" + f
        f = f.upper()
        f = f.replace("{ACC_RATE}", str(a_rate))
        f = f.replace("{PROTOCOL_RATE}", str(p_rate))
        f = f.replace("{PROJECT_COEFF_RATE}", str(p_coeff_rate))
        f = f.replace("{VAT_RATE}", str(v_rate))
        f = f.replace("{STAMP_TAX_RATE}", str(stamp_tax_rate))
        f = f.replace("{TURKAK_RATE}", str(turkak_rate))
        return f

    safe_write_logic('C', inject_rates(f_c), "TOPLAM PROTOKOLLÜ HİZMET BEDELİ")
    safe_write_logic('D', inject_rates(f_d), f"KDV (%{v_rate})")
    safe_write_logic('E', inject_rates(f_e), f"DAMGA VERGİSİ (‰{stamp_tax_rate*1000:.2f})")
    safe_write_logic('F', inject_rates(f_f), f"TÜRKAK PAYI (A * %{turkak_rate})")
    safe_write_logic('G', inject_rates(f_g), "GENEL TOPLAM")
    
    # WRITE SIGNATURES (Footer)
    # Düzenleyen (28,2), Kontrol Eden (28,7), Onaylayan (28,10)
    # Row 29: Names, Row 30: Titles
    sig_font = Font(bold=True)
    def write_sig(row_idx, col, val):
        if not val: return
        cell = ws.cell(row=row_idx, column=col)
        if isinstance(cell, openpyxl.cell.cell.MergedCell):
             for m_range in ws.merged_cells.ranges:
                if cell.coordinate in m_range:
                    master = ws.cell(row=m_range.min_row, column=m_range.min_col)
                    master.value = val; master.font = sig_font if row_idx == 29 else None
        else: 
            cell.value = val
            if row_idx == 29: cell.font = sig_font

    write_sig(29, 2, sig_d)  # Column B
    write_sig(29, 7, sig_k)  # Column G
    write_sig(29, 10, sig_o) # Column J

    write_sig(30, 2, sig_d_uv)
    write_sig(30, 7, sig_k_uv)
    write_sig(30, 10, sig_o_uv)

    # Save
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
