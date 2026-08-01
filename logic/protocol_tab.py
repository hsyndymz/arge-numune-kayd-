import streamlit as st
from logic.kesif_tab import calculate_project_costs
import pandas as pd
import sqlite3
import os
import io
from datetime import datetime
from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from logic.db_utils import get_connection as get_data_connection, get_settings_connection

# Redirect protocol records to settings DB, but keep project lookups in Data DB
def get_connection():
    return get_settings_connection()

# --- Helper: Turkish Character Fix ---
def tr_fix(text):
    if not text: return ""
    replacements = {'ı': 'i', 'İ': 'I', 'ğ': 'g', 'Ğ': 'G', 'ü': 'u', 'Ü': 'U', 'ş': 's', 'Ş': 'S', 'ö': 'o', 'Ö': 'O', 'ç': 'c', 'Ç': 'C'}
    for k, v in replacements.items():
        text = text.replace(k, v)
    return text

# --- CRUD Operations ---
def get_active_protocols(month=None, firm_filter=None):
    conn = get_connection()
    query = "SELECT * FROM protocols WHERE is_archived=0"
    params = []
    
    if month:
        query += " AND month = ?"
        params.append(month)
        
    if firm_filter:
        query += " AND firm LIKE ?"
        params.append(f"%{firm_filter}%")
        
    query += " ORDER BY payment_date DESC, id DESC"
    
    df = pd.read_sql(query, conn, params=params)
    conn.close()
    return df

def get_archived_protocols(year):
    conn = get_connection()
    df = pd.read_sql("SELECT * FROM protocols WHERE is_archived=1 AND archive_year=?", conn, params=(str(year),))
    conn.close()
    return df

def add_protocol(data):
    conn = get_connection()
    cursor = conn.cursor()
    
    # Calculate totals if not provided (though mapped from UI)
    base = float(data.get('base_cost', 0.0))
    kdv = float(data.get('kdv_amount', 0.0))
    stamp = float(data.get('stamp_tax', 0.0))
    turkak = float(data.get('turkak_fee', 0.0))
    
    # Auto-calculate total if 0
    if float(data.get('total_amount', 0.0)) == 0.0:
        total = base + kdv + stamp + turkak
    else:
        total = float(data.get('total_amount', 0.0))

    # Helper to standardize date to YYYY-MM-DD
    def standardize_date(d_str):
        if not d_str: return None
        try:
            # Assume input is DD.MM.YYYY or similar
            if '.' in d_str:
                parts = d_str.split('.')
                if len(parts) == 3:
                    return f"{parts[2]}-{parts[1]}-{parts[0]}"
            return d_str
        except:
            return d_str

    p_date = standardize_date(data.get('protocol_date'))
    pay_date = standardize_date(data.get('payment_date')) or p_date

    cursor.execute("""
        INSERT INTO protocols (
            sequence_no, office_record_no, protocol_no, region_no, sender, firm, job_description, 
            protocol_date, base_cost, kdv_amount, total_cost_with_kdv, turkak_fee, secondary_keşif_with_kdv, 
            stamp_tax, total_amount, payment_date, receipt_no, bank_info, month, linked_project_name
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        data.get('sequence_no'), data.get('office_record_no'), data.get('protocol_no'), data.get('region_no'),
        data.get('sender'), data.get('firm'), data.get('job_description'), p_date,
        base, kdv, data.get('total_cost_with_kdv'), turkak, data.get('secondary_keşif_with_kdv'),
        stamp, total, pay_date, data.get('receipt_no'), data.get('bank_info'), 
        data.get('month'), data.get('linked_project_name')
    ))
    conn.commit()
    conn.close()

def update_protocol(p_id, data):
    conn = get_connection()
    cursor = conn.cursor()
    
    # Helper to construct UPDATE query dynamically
    fields = []
    values = []
    for k, v in data.items():
        fields.append(f"{k}=?")
        values.append(v)
    
    values.append(p_id)
    query = f"UPDATE protocols SET {', '.join(fields)} WHERE id=?"
    
    cursor.execute(query, values)
    conn.commit()
    conn.close()

def delete_protocol(p_id):
    conn = get_connection()
    conn.execute("DELETE FROM protocols WHERE id=?", (p_id,))
    conn.commit()
    conn.close()

def archive_year_protocols(year):
    conn = get_connection()
    # Archive where payment_date starts with year
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE protocols SET is_archived=1, archive_year=? 
        WHERE is_archived=0 AND payment_date LIKE ?
    """, (str(year), f"{year}-%"))
    count = cursor.rowcount
    conn.commit()
    conn.close()
    return count

# --- PDF Generation ---
def generate_protocol_pdf(protocol_row):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)
    
    # Font setup (standard fonts)
    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 10, tr_fix("PROTOKOL RAPORU"), ln=True, align="C")
    pdf.ln(10)
    
    try:
        # Logo if exists
        if os.path.exists("logo.jpg"):
            pdf.image("logo.jpg", 10, 8, 33)
    except: pass

    fields = [
        ("Protokol No", protocol_row['protocol_no']),
        ("Firma", protocol_row['firm']),
        ("Is Tanimi", protocol_row['job_description']),
        ("Tarih", protocol_row['protocol_date']),
        ("Toplam Tutar", f"{protocol_row['total_amount']:.2f} TL"),
        ("Banka Bilgisi", protocol_row['bank_info']),
        ("Odeme Tarihi", protocol_row['payment_date']),
        ("Dekont No", protocol_row['receipt_no'])
    ]
    
    for label, value in fields:
        pdf.set_font("Arial", "B", 12)
        pdf.cell(50, 10, tr_fix(label) + ": ", 0)
        pdf.set_font("Arial", "", 12)
        pdf.multi_cell(0, 10, tr_fix(str(value)) if value else "-", border=0)
        pdf.ln(2)

    # Footer signature area
    pdf.ln(30)
    y = pdf.get_y()
    pdf.set_font("Arial", "B", 10)
    pdf.cell(60, 5, tr_fix("Hazirlayan"), align='C')
    pdf.cell(60, 5, tr_fix("Kontrol Eden"), align='C')
    pdf.cell(60, 5, tr_fix("Onaylayan"), align='C')
    
    filename = f"protokol_{protocol_row['id']}_{datetime.now().strftime('%Y%m%d')}.pdf"
    
    # Return as bytes
    return pdf.output(dest='S').encode('latin-1')

# --- Excel Export Logic ---
def export_simple_annual_summary(year):
    """Export simple annual summary (Protocol count + Total amount) matching template 2 format"""
    conn = get_connection()
    df = pd.read_sql("SELECT * FROM protocols WHERE payment_date LIKE ?", conn, params=(f"{year}-%",))
    conn.close()
    
    months_tr = {
        "01": "OCAK", "02": "ŞUBAT", "03": "MART", "04": "NİSAN", "05": "MAYIS", "06": "HAZİRAN",
        "07": "TEMMUZ", "08": "AĞUSTOS", "09": "EYLÜL", "10": "EKİM", "11": "KASIM", "12": "ARALIK"
    }
    
    summary_data = {m: {"count": 0, "total": 0} for m in months_tr}
    
    for _, row in df.iterrows():
        if row['payment_date'] and len(row['payment_date']) >= 7:
            m_code = row['payment_date'].split('-')[1]
            if m_code in summary_data:
                summary_data[m_code]['count'] += 1
                summary_data[m_code]['total'] += row['total_amount'] or 0
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{year} Protokol Özet"
    
    # Header
    ws.merge_cells('A1:D1')
    ws['A1'] = f"01.01-31.12.{year} TARİHLERİ ARASINDA"
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].font = Font(bold=True, size=12)
    
    ws.merge_cells('A2:D2')
    ws['A2'] = "KARAYOLLARI 9. BÖLGE MÜDÜRLÜĞÜ"
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['A2'].font = Font(bold=True)
    
    ws.merge_cells('A3:D3')
    ws['A3'] = "ARAŞTIRMA VE GELİŞTİRME BAŞMÜHENDİSLİĞİ LAB.YAPILAN DENEY"
    ws['A3'].alignment = Alignment(horizontal='center')
    ws['A3'].font = Font(bold=True)
    
    ws.merge_cells('A4:D4')
    ws['A4'] = "PROTOKOL LİSTESİ"
    ws['A4'].alignment = Alignment(horizontal='center')
    ws['A4'].font = Font(bold=True)
    
    # Column headers
    border = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )
    
    headers = ["AY", "PROTOKOL SAYISI", "PARA MİKTARI\n(TL)", "NOT"]
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Data rows
    row_num = 7
    total_count = 0
    total_amount = 0
    
    for m in sorted(months_tr.keys()):
        d = summary_data[m]
        ws.cell(row=row_num, column=1, value=months_tr[m])
        ws.cell(row=row_num, column=2, value=d['count'])
        ws.cell(row=row_num, column=3, value=d['total']).number_format = '#,##0.00'
        ws.cell(row=row_num, column=4, value='')
        
        for col in range(1, 5):
            ws.cell(row=row_num, column=col).border = border
        
        total_count += d['count']
        total_amount += d['total']
        row_num += 1
    
    # Total row
    ws.cell(row=row_num, column=1, value="TOPLAM").font = Font(bold=True)
    ws.cell(row=row_num, column=2, value=total_count).font = Font(bold=True)
    ws.cell(row=row_num, column=3, value=total_amount).font = Font(bold=True)
    ws.cell(row=row_num, column=3).number_format = '#,##0.00'
    
    for col in range(1, 5):
        ws.cell(row=row_num, column=col).border = border
    
    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 30
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def export_annual_detailed_excel(year):
    """Export annual summary by month matching template 3 format"""
    conn = get_connection()
    df = pd.read_sql("SELECT * FROM protocols WHERE payment_date LIKE ?", conn, params=(f"{year}-%",))
    conn.close()
    
    months_tr = {
        "01": "Ocak", "02": "Şubat", "03": "Mart", "04": "Nisan", "05": "Mayıs", "06": "Haziran",
        "07": "Temmuz", "08": "Ağustos", "09": "Eylül", "10": "Ekim", "11": "Kasım", "12": "Aralık"
    }
    
    # Group by Month
    summary_data = {m: {"count": 0, "base": 0, "kdv": 0, "stamp": 0, "turkak": 0, "total": 0} for m in months_tr}
    
    for _, row in df.iterrows():
        if row['payment_date'] and len(row['payment_date']) >= 7:
            m_code = row['payment_date'].split('-')[1] # YYYY-MM-DD
            if m_code in summary_data:
                d = summary_data[m_code]
                d['count'] += 1
                d['base'] += row['base_cost'] or 0
                d['kdv'] += row['kdv_amount'] or 0
                d['stamp'] += row['stamp_tax'] or 0
                d['turkak'] += row['turkak_fee'] or 0
                d['total'] += row['total_amount'] or 0
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{year} Yıl Deney Ücretleri"
    
    # Header Section
    peach_fill = PatternFill(start_color="FFD9B3", end_color="FFD9B3", fill_type="solid")
    
    ws.merge_cells('A1:G1')
    ws['A1'] = "Karayolları 9. Bölge Müdürlüğü"
    ws['A1'].fill = peach_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].font = Font(bold=True)
    
    ws.merge_cells('A2:G2')
    ws['A2'] = "(Araştırma ve Geliştirme Başmühendisliği)"
    ws['A2'].fill = peach_fill
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A3:G3')
    ws['A3'] = f"01 Ocak {year}-31 Aralık {year} Yılı Deney Ücretleri"
    ws['A3'].fill = peach_fill
    ws['A3'].alignment = Alignment(horizontal='center')
    
    # Column Headers (Row 5)
    headers = ["Ay", "Protokol Adedi", "Protokollü Hizmet Bedeli", "KDV", "Damga Vergisi", "Türkak Payı", "Genel Toplam"]
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=idx, value=header)
        cell.fill = peach_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Data Rows
    totals = {"count": 0, "base": 0, "kdv": 0, "stamp": 0, "turkak": 0, "total": 0}
    row_num = 6
    
    for m in sorted(months_tr.keys()):
        d = summary_data[m]
        data_row = [
            months_tr[m], 
            d['count'], 
            d['base'], 
            d['kdv'], 
            d['stamp'], 
            d['turkak'], 
            d['total']
        ]
        
        for col_idx, value in enumerate(data_row, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.border = border
            
            # Currency formatting
            if col_idx >= 3:
                cell.number_format = '#,##0.00 ₺'
        
        totals['count'] += d['count']
        totals['base'] += d['base']
        totals['kdv'] += d['kdv']
        totals['stamp'] += d['stamp']
        totals['turkak'] += d['turkak']
        totals['total'] += d['total']
        
        row_num += 1
    
    # Totals Row
    total_row = ["Toplamlar", totals['count'], totals['base'], totals['kdv'], totals['stamp'], totals['turkak'], totals['total']]
    for col_idx, value in enumerate(total_row, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.fill = peach_fill
        cell.font = Font(bold=True)
        cell.border = border
        if col_idx >= 3:
            cell.number_format = '#,##0.00 ₺'
    
    # Footer (Signatures)
    ws.merge_cells(f'A{row_num+2}:B{row_num+2}')
    ws[f'A{row_num+2}'] = "Hazırlayan"
    ws[f'A{row_num+2}'].font = Font(bold=True)
    
    ws.merge_cells(f'C{row_num+2}:D{row_num+2}')
    ws[f'C{row_num+2}'] = "Kontrol Eden"
    ws[f'C{row_num+2}'].font = Font(bold=True)
    
    ws.merge_cells(f'F{row_num+2}:G{row_num+2}')
    ws[f'F{row_num+2}'] = "Onaylayan"
    ws[f'F{row_num+2}'].font = Font(bold=True)
    
    ws[f'A{row_num+3}'] = "Sibel GÜMÜŞ"
    ws[f'C{row_num+3}'] = "Jiyan KARAKAŞ"
    ws[f'F{row_num+3}'] = "Atakan ERSOY"
    
    ws[f'A{row_num+4}'] = "Büro ve Kayıt Memuru"
    ws[f'C{row_num+4}'] = "Üst.yapı Gel.Müh"
    ws[f'F{row_num+4}'] = "Ar-Ge Başmühendisi"
    
    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def export_full_list_excel(month_filter=None):
    """Export detailed monthly protocol list matching the specific 18-column technical template"""
    df = get_active_protocols(month=month_filter)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PROTOKOL LİSTESİ"
    
    # --- STYLE DEFINITIONS ---
    peach_fill = PatternFill(start_color="FFD9B3", end_color="FFD9B3", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    border_medium = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))
    
    center_aligned = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_aligned = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # --- HEADER SECTION (Rows 1-5) ---
    ws.merge_cells('A1:R1')
    ws['A1'] = "PROTOKOL LİSTESİ"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center_aligned
    
    ws.merge_cells('A2:R2')
    ws['A2'] = "Karayolları 9. Bölge Müdürlüğü"
    ws['A2'].alignment = center_aligned
    ws['A2'].font = Font(bold=True)
    
    if month_filter:
        year_month = month_filter.split('-')
        month_name = {
            "01": "OCAK", "02": "ŞUBAT", "03": "MART", "04": "NİSAN", "05": "MAYIS", "06": "HAZİRAN",
            "07": "TEMMUZ", "08": "AĞUSTOS", "09": "EYLÜL", "10": "EKİM", "11": "KASIM", "12": "ARALIK"
        }.get(year_month[1], "")
        ws.merge_cells('A3:R3')
        ws['A3'] = f"{year_month[0]} YILI {month_name} AYI"
        ws['A3'].alignment = center_aligned
        ws['A3'].font = Font(bold=True)
    
    ws.merge_cells('A4:R4')
    ws['A4'] = "AR-GE Dairesi Bşk. Hizmetleri Protokollü İşler Listesi"
    ws['A4'].alignment = center_aligned

    # --- MAIN HEADERS (Row 6) ---
    headers = [
        "Sıra\nNo:", "Büro\nKayıt No:", "Protokol\nNo:", "Bölge\nNo:", "NUMUNEYİ GÖNDEREN", "PROTOKOL\nBEDELİNİ\nYATIRAN FİRMA", 
        "YAPILACAK İŞ", "Protokol\nİmzalanma\nTarihi", "Deney\n1.Keşif Özeti\n(TL)", "K.D.V.\n( H * %20 )\n(TL)", 
        "KDV'li \n1.Keşif Özeti\n(TL)", "K.D.V.\n2.Keşif Tutarı\n(TL)", "TÜRKAK\nPayı\n( G * %1 )", "Damga\nVergisi", 
        "Toplam\nMiktar\n( TL )", "DEKONT\nTarihi", "DEKONT NO:", "AÇIKLAMA"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.fill = yellow_fill
        cell.font = Font(bold=True, size=8)
        cell.alignment = center_aligned
        cell.border = border_medium

    # --- SUB-LABELS (Row 7) ---
    sub_labels = {9: "H", 10: "I", 11: "J", 12: "İ", 13: "K", 15: "H+K+(İ)+(I)"}
    for col in range(1, 19):
        cell = ws.cell(row=7, column=col, value=sub_labels.get(col, ""))
        cell.fill = yellow_fill
        cell.font = Font(bold=True, size=8)
        cell.alignment = center_aligned
        cell.border = border_medium

    # --- DATA ROWS ---
    row_num = 8
    for idx, row_data in df.iterrows():
        # Mapping DB fields to Template cols
        data_row = [
            row_data.get('sequence_no', ''), # A
            row_data.get('office_record_no', ''), # B
            row_data.get('protocol_no', ''), # C
            row_data.get('region_no', '9'), # D
            row_data.get('sender', ''), # E
            row_data.get('firm', ''), # F
            row_data.get('job_description', ''), # G
            row_data.get('protocol_date', ''), # H
            row_data.get('base_cost', 0), # I (H in template)
            row_data.get('kdv_amount', 0), # J (I in template)
            row_data.get('total_cost_with_kdv', 0), # K (J in template)
            row_data.get('secondary_keşif_with_kdv', 0), # L (İ in template)
            row_data.get('turkak_fee', 0), # M (K in template)
            row_data.get('stamp_tax', 0), # N
            row_data.get('total_amount', 0), # O
            row_data.get('payment_date', ''), # P
            row_data.get('receipt_no', ''), # Q
            row_data.get('bank_info', '') # R
        ]
        
        for col_idx, value in enumerate(data_row, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.border = border_thin
            cell.alignment = center_aligned if col_idx != 5 and col_idx != 6 and col_idx != 7 else left_aligned
            
            # Formatting Currency
            if col_idx in range(9, 16):
                cell.number_format = '#,##0.00'
        
        row_num += 1

    # --- TOTALS ROW ---
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=8)
    total_label = ws.cell(row=row_num, column=1, value="TOPLAM")
    total_label.alignment = Alignment(horizontal='right')
    total_label.font = Font(bold=True)
    
    sum_cols = [9, 10, 11, 12, 14, 15] # H, I, J, İ, Damga, Toplam
    for col_idx in range(1, 19):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.fill = peach_fill
        cell.border = border_medium
        if col_idx in sum_cols:
            col_letter = get_column_letter(col_idx)
            cell.value = f"=SUM({col_letter}8:{col_letter}{row_num-1})"
            cell.number_format = '#,##0.00'
            cell.font = Font(bold=True, color="FF0000")

    row_num += 2

    # --- FOOTER (SIGNATURES & NOTES) ---
    # Notes Section
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=10)
    ws.cell(row=row_num, column=1, value="NOT 1: H, I, İ, G, J Keşif Özetinde Verilen Miktarlar Olup Keşif Özetinden Alınacaktır.").font = Font(size=8)
    row_num += 1
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=10)
    ws.cell(row=row_num, column=1, value="NOT 2: Toplam Miktar hesaplanırken ( İ ) veya ( İ )'den Biri Alınacak.").font = Font(size=8)
    
    # Signature Section (Starts around col 11)
    sig_start_row = row_num - 1
    ws.cell(row=sig_start_row, column=11, value="Hazırlayan").border = border_thin
    ws.cell(row=sig_start_row+1, column=11, value="Adı-Soyadı").border = border_thin
    ws.cell(row=sig_start_row+2, column=11, value="Ünvanı").border = border_thin
    ws.cell(row=sig_start_row+3, column=11, value="İmza").border = border_thin
    
    ws.merge_cells(start_row=sig_start_row, start_column=12, end_row=sig_start_row, end_column=18)
    ws.cell(row=sig_start_row, column=12, value="Sibel GÜMÜŞ").border = border_thin
    
    ws.merge_cells(start_row=sig_start_row+1, start_column=12, end_row=sig_start_row+1, end_column=18)
    ws.cell(row=sig_start_row+1, column=12, value="Atakan ERSOY").border = border_thin
    
    ws.merge_cells(start_row=sig_start_row+2, start_column=12, end_row=sig_start_row+2, end_column=18)
    ws.cell(row=sig_start_row+2, column=12, value="Başmühendis").border = border_thin
    
    ws.merge_cells(start_row=sig_start_row+3, start_column=12, end_row=sig_start_row+3, end_column=18)
    ws.cell(row=sig_start_row+3, column=12, value="").border = border_thin

    # Formatting columns widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 6
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 35
    ws.column_dimensions['H'].width = 12
    for c in "IJKLMNO": ws.column_dimensions[c].width = 10
    ws.column_dimensions['P'].width = 12
    ws.column_dimensions['Q'].width = 12
    ws.column_dimensions['R'].width = 20
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# --- Data Import ---
def search_col(df, possible_names):
    for col in df.columns:
        if str(col).strip().replace('\n', '') in possible_names:
            return col
        # Fuzzy match attempt
        norm_col = str(col).lower().replace(' ', '').replace('\n', '')
        for p in possible_names:
            norm_p = p.lower().replace(' ', '').replace('\n', '')
            if norm_p in norm_col:
                return col
    return None

def import_excel_data(file):
    try:
        xl = pd.ExcelFile(file)
        conn = get_connection()
        cursor = conn.cursor()
        
        imported_count = 0
        
        for sheet in xl.sheet_names:
            df = pd.read_excel(file, sheet_name=sheet, skiprows=4) # Assuming format
            
            # Map columns
            # Firm, Work, Date, TotalAmount are critical
            firm_col = search_col(df, ["PROTOKOL BEDELİNİ YATIRAN FİRMA", "FİRMA", "Firm"])
            work_col = search_col(df, ["YAPILACAK İŞ", "İŞİN ADI", "Work"])
            total_col = search_col(df, ["Toplam Miktar ( TL )", "Toplam Miktar", "Genel Toplam"])
            date_col = search_col(df, ["Protokol İmzalanma Tarihi", "Tarih"])
            
            # Default fallback if names not found (using indices like in original script)
            if not firm_col and len(df.columns) > 8: firm_col = df.columns[8]
            
            for _, row in df.iterrows():
                # Basic check for valid row
                firm_val = row[firm_col] if firm_col and pd.notna(row[firm_col]) else None
                
                if firm_val:
                    # Clean data
                    def get_v(c): return row[c] if c and c in row and pd.notna(row[c]) else None
                    
                    cursor.execute("""
                        INSERT INTO protocols (firm, job_description, total_amount, protocol_date, month, is_archived)
                        VALUES (?, ?, ?, ?, ?, 0)
                    """, (str(firm_val), str(get_v(work_col)), float(get_v(total_col)) if get_v(total_col) else 0.0, 
                          str(get_v(date_col)), sheet))
                    imported_count += 1
        
        conn.commit()
        conn.close()
        return True, f"{imported_count} kayıt başarıyla aktarıldı."
    except Exception as e:
        return False, f"Hata: {str(e)}"

# --- MAIN RENDER FUNCTION ---
def render_protocol_tab():
    st.title("📑 Protokol ve Maliyet Takip Sistemi")
    
    # Top Menu
    menu = st.radio("İşlemler", ["📋 Protokol Listesi", "➕ Yeni Protokol Ekle", "📤 Excel'den Aktar", "📊 Raporlar ve İstatistikler"], horizontal=True)
    
    if menu == "📋 Protokol Listesi":
        st.subheader("Aktif Protokoller")
        
        c1, c2 = st.columns([3, 1])
        with c1:
            search = st.text_input("🔍 Firma veya İş Tanımı Ara", key="prot_search")
        with c2:
            # Dynamic month filter from database
            try:
                available_months = pd.read_sql("SELECT DISTINCT month FROM protocols WHERE month IS NOT NULL ORDER BY month DESC", get_connection())['month'].tolist()
            except:
                available_months = []
                
            month_filter = st.selectbox("📅 İşlem Dönemi Filtresi", ["Tümü"] + available_months)
        
        df = get_active_protocols(
            month=None if month_filter == "Tümü" else month_filter,
            firm_filter=search
        )
        
        st.dataframe(df, use_container_width=True)
        
        # CRUD Actions
        with st.expander("✏️ Seçili Protokolü Düzenle / Sil"):
            if not df.empty:
                sel_id = st.selectbox("Düzenlenecek Protokol ID", df['id'].tolist(), format_func=lambda x: f"ID: {x} - {df[df['id']==x]['firm'].values[0]}")
                row = df[df['id'] == sel_id].iloc[0]
                
                with st.form("edit_prot_form"):
                    c_1, c_2 = st.columns(2)
                    u_firm = c_1.text_input("Firma", value=row['firm'])
                    u_desc = c_2.text_area("İş Tanımı", value=row['job_description'])
                    u_date = c_1.text_input("Tarih (YYYY-MM-DD)", value=row['protocol_date'])
                    
                    c_3, c_4 = st.columns(2)
                    u_total = c_3.number_input("Toplam Tutar", value=float(row['total_amount'] or 0.0))
                    u_receipt = c_4.text_input("Dekont No", value=row['receipt_no'] if row['receipt_no'] else "")
                    
                    c_5, c_6 = st.columns(2)
                    u_pay_date = c_5.text_input("Ödeme Tarihi (YYYY-MM-DD)", value=row['payment_date'] if row['payment_date'] else "")
                    
                    # İşlem Dönemi Edit Field
                    curr_m = row['month'] if row['month'] else datetime.now().strftime("%Y-%m")
                    u_month = c_6.text_input("İşlem Dönemi (YYYY-MM)", value=curr_m)
                    
                    u_bank = st.text_input("Açıklama / Banka", value=row['bank_info'] if row['bank_info'] else "")
                    
                    if st.form_submit_button("GÜNCELLE"):
                        update_protocol(sel_id, {
                            "firm": u_firm, "job_description": u_desc, 
                            "protocol_date": u_date, "total_amount": u_total,
                            "receipt_no": u_receipt, "payment_date": u_pay_date, 
                            "month": u_month, "bank_info": u_bank
                        })
                        st.success("Güncellendi!")
                        st.rerun()
                
                if st.button("🗑️ BU PROTOKOLÜ SİL", type="primary"):
                    delete_protocol(sel_id)
                    st.success("Silindi!")
                    st.rerun()
            else:
                st.info("Listelenecek protokol yok.")

    elif menu == "➕ Yeni Protokol Ekle":
        st.subheader("Yeni Protokol Oluştur")
        
        # Project Linking Feature - MUST USE DATA DB for samples table
        conn_data = get_data_connection()
        projects = pd.read_sql("SELECT DISTINCT proje, firma FROM samples WHERE proje IS NOT NULL AND proje != ''", conn_data)
        conn_data.close()
        
        proj_list = [f"{r['proje']} | {r['firma']}" for _, r in projects.iterrows()]
        link_choice = st.selectbox("🔗 Mevcut Projeden Bilgi Çek (İsteğe Bağlı)", ["Bağlantı Yok"] + proj_list)
        
        default_firm = ""
        default_job = ""
        linked_proj = ""
        
        if link_choice != "Bağlantı Yok":
            linked_proj = link_choice.split(" | ")[0]
            default_firm = link_choice.split(" | ")[1]
            default_job = f"{linked_proj} Projesi Kapsamındaki Deneyler"
        
        # Auto-calculate defaults from Kesif if project linked
        def_base = 0.0
        def_kdv = 0.0
        def_stamp = 0.0
        def_turkak = 0.0
        
        if linked_proj:
            costs = calculate_project_costs(linked_proj)
            # Check if we actually found distinct costs (Total > 0 implies data exists)
            if costs['total_amount'] > 0:
                def_base = costs['base_cost']
                def_kdv = costs['kdv_amount']
                def_stamp = costs['stamp_tax']
                def_turkak = costs['turkak_fee']
                st.success(f"✅ '{linked_proj}' projesi için keşif verileri bulundu ve aşağıya yansıtıldı. Gerekirse müdahale edebilirsiniz.")
            else:
                st.warning(f"⚠️ '{linked_proj}' projesi için kayıtlı bir keşif hesaplaması bulunamadı veya tutar 0. Lütfen maliyetleri manuel giriniz.")

        with st.form("new_prot_form"):
            c1, c2 = st.columns(2)
            firm = c1.text_input("Firma Ünvanı", value=default_firm)
            job = c2.text_area("Yapılacak İş", value=default_job)
            
            c3, c4, c5 = st.columns(3)
            p_date = c3.date_input("Protokol Tarihi")
            pay_date = c4.date_input("Ödeme Tarihi")
            receipt = c5.text_input("Dekont No")

            st.write("---")
            # New: Explicit Month Selection for Accounting Config
            # Default to pay_date's month
            default_m_str = pay_date.strftime("%Y-%m")
            m_list = []
            
            # Generate a list of last 12 months + next 2 months
            import datetime
            curr = datetime.date.today().replace(day=1)
            start_date = curr - datetime.timedelta(days=365)
            for i in range(15):
                d = start_date + datetime.timedelta(days=30*i)
                m_str = d.strftime("%Y-%m")
                if m_str not in m_list: m_list.append(m_str)
            m_list.sort(reverse=True)
            
            try:
                def_idx = m_list.index(default_m_str)
            except:
                def_idx = 0

            selected_month = st.selectbox("İşlem Dönemi (Hangi aya dahil edilecek?)", m_list, index=def_idx)
            
            st.markdown("---")
            st.markdown("**Maliyet Detayları**")
            m1, m2, m3, m4 = st.columns(4)
            base = m1.number_input("Hizmet Bedeli (Taban)", min_value=0.0, value=def_base, step=0.01)
            kdv = m2.number_input("KDV", min_value=0.0, value=def_kdv, step=0.01)
            stamp = m3.number_input("Damga Vergisi", min_value=0.0, value=def_stamp, step=0.01)
            turkak = m4.number_input("TÜRKAK Payı", min_value=0.0, value=def_turkak, step=0.01)
            
            total_calc = base + kdv + stamp + turkak
            st.info(f"Hesaplanan Toplam: {total_calc:,.2f} TL")
            
            if st.form_submit_button("KAYDET", type="primary"):
                data = {
                    "firm": firm, "job_description": job, 
                    "protocol_date": p_date.strftime("%Y-%m-%d"),
                    "payment_date": pay_date.strftime("%Y-%m-%d"),
                    "receipt_no": receipt,
                    "month": selected_month, # Explicitly save selected month
                    "base_cost": base, "kdv_amount": kdv, 
                    "stamp_tax": stamp, "turkak_fee": turkak,
                    "total_amount": total_calc,
                    "linked_project_name": linked_proj
                }
                add_protocol(data)
                st.success("✅ Protokol başarıyla kaydedildi!")
    
    elif menu == "📤 Excel'den Aktar":
        st.subheader("Toplu Veri Aktarımı")
        st.info("Sibel Hanım'ın formatındaki Excel dosyasını buradan yükleyebilirsiniz.")
        
        up_file = st.file_uploader("Excel Dosyası Seç (.xlsx)", type=['xlsx'])
        if up_file:
            if st.button("VERİLERİ İÇERİ AKTAR"):
                success, msg = import_excel_data(up_file)
                if success:
                    st.success(msg)
                else:
                    st.error(msg)
                    
    elif menu == "📊 Raporlar ve İstatistikler":
        st.subheader("Rapor Merkezi")
        
        c1, c2 = st.columns(2)
        with c1:
            st.markdown("#### 📥 Excel Çıktıları")
            year_sel = st.number_input("Yıl Seçin", value=2026, step=1)
            
            # Simple Annual Summary
            simple_excel = export_simple_annual_summary(year_sel)
            st.download_button("📥 Basit Yıllık Özet (Sayı + Tutar)", simple_excel, f"Basit_Ozet_{year_sel}.xlsx")
            
            st.markdown("---")
            # Detailed Annual Report
            detailed_excel = export_annual_detailed_excel(year_sel)
            st.download_button("📥 Detaylı Yıllık Rapor (Tüm Kolon)", detailed_excel, f"Yillik_Detayli_{year_sel}.xlsx")
            
            st.markdown("---")
            st.markdown("**Aylık Rapor**")
            try:
                available_months_report = pd.read_sql("SELECT DISTINCT month FROM protocols WHERE month IS NOT NULL ORDER BY month DESC", get_connection())['month'].tolist()
            except:
                available_months_report = []
                
            month_sel = st.selectbox("İşlem Dönemi Seç", ["Tümü"] + available_months_report, key="monthly_report_sel")
            month_param = None if month_sel == "Tümü" else month_sel
            monthly_list = export_full_list_excel(month_filter=month_param)
            st.download_button("📥 Aylık Protokol Listesi (Excel)", monthly_list, f"Protokoller_{month_sel}.xlsx")
            
            st.markdown("---")
            full_list = export_full_list_excel()
            st.download_button("📥 Tüm Liste (Excel)", full_list, "Tum_Protokoller.xlsx")
            
        with c2:
            st.markdown("#### 📈 Özet Grafikler")
            df = get_active_protocols()
            if not df.empty:
                df['month_key'] = df['payment_date'].apply(lambda x: x[:7] if x else "N/A")
                monthly_rev = df.groupby('month_key')['total_amount'].sum()
                st.bar_chart(monthly_rev)
                st.caption("Aylık Toplam Ciro")
