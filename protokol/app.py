from fastapi import FastAPI, Depends, HTTPException
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy.orm import Session
from database import SessionLocal, Protocol
from pydantic import BaseModel
from typing import List, Optional
import os
import io
import pandas as pd
from fpdf import FPDF
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

app = FastAPI()

# Pydantic models for API
class ProtocolBase(BaseModel):
    sequence_no: Optional[int] = None
    office_record_no: Optional[str] = None
    protocol_no: Optional[str] = None
    region_no: Optional[int] = None
    sender: Optional[str] = None
    firm: Optional[str] = None
    job_description: Optional[str] = None
    protocol_date: Optional[str] = None
    base_cost: Optional[float] = 0.0
    kdv_amount: Optional[float] = 0.0
    total_cost_with_kdv: Optional[float] = 0.0
    turkak_fee: Optional[float] = 0.0
    secondary_keşif_with_kdv: Optional[float] = 0.0
    stamp_tax: Optional[float] = 0.0
    total_amount: Optional[float] = 0.0
    payment_date: Optional[str] = None
    receipt_no: Optional[str] = None
    bank_info: Optional[str] = None
    month: Optional[str] = None

class ProtocolSchema(ProtocolBase):
    id: int

    class Config:
        from_attributes = True

# Dependency to get DB session
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# API Endpoints
@app.get("/api/protocols", response_model=List[ProtocolSchema])
def read_protocols(skip: int = 0, limit: int = 100, firm: Optional[str] = None, month: Optional[str] = None, include_archived: bool = False, db: Session = Depends(get_db)):
    query = db.query(Protocol)
    if not include_archived:
        query = query.filter(Protocol.is_archived == 0)
    
    if firm:
        query = query.filter(Protocol.firm.contains(firm))
    if month:
        # UPDATED: Filter by the explicit 'month' column (YYYY-MM)
        # This allows handling late receipts in the current processing period.
        query = query.filter(Protocol.month == month)
    return query.order_by(Protocol.payment_date.desc()).offset(skip).limit(limit).all()

@app.get("/api/protocols/archived-years")
def get_archived_years(db: Session = Depends(get_db)):
    years = db.query(Protocol.archive_year).filter(Protocol.is_archived == 1).distinct().all()
    return [y[0] for y in years if y[0]]

@app.get("/api/protocols/archive/{year}", response_model=List[ProtocolSchema])
def read_archived_protocols(year: str, db: Session = Depends(get_db)):
    return db.query(Protocol).filter(Protocol.is_archived == 1, Protocol.archive_year == year).all()

@app.get("/api/protocols/active-months")
def get_active_months(db: Session = Depends(get_db)):
    # Get distinct payment dates from active protocols
    dates = db.query(Protocol.payment_date).filter(Protocol.is_archived == 0).distinct().all()
    # Extract YYYY-MM
    months = set()
    for (d,) in dates:
        if d and len(d) >= 7:
            months.add(d[:7])
    return sorted(list(months), reverse=True)

@app.post("/api/protocols/close-year")
def close_year(year: str, db: Session = Depends(get_db)):
    # Archive all protocols where payment_date is in that year and is_archived is 0
    protocols = db.query(Protocol).filter(Protocol.is_archived == 0, Protocol.payment_date.like(f"{year}-%")).all()
    if not protocols:
        raise HTTPException(status_code=404, detail="Bu yıla ait kapatılacak kayıt bulunamadı.")
    
    for p in protocols:
        p.is_archived = 1
        p.archive_year = year
    
    db.commit()
    return {"message": f"{year} yılı başarıyla kapatıldı ve {len(protocols)} kayıt arşivlendi."}

@app.get("/api/reports/annual-detailed/{year}")
def export_annual_detailed(year: str, db: Session = Depends(get_db)):
    # Resim 1 Formatı: Ay, Protokol Adedi, Protokollü Hizmet Bedeli, KDV, Damga Vergisi, Türkak Payı, Genel Toplam
    # Using payment_date for grouping by months
    protocols = db.query(Protocol).filter(Protocol.payment_date.like(f"{year}-%")).all()
    
    months_tr = {
        "01": "Ocak", "02": "Şubat", "03": "Mart", "04": "Nisan", "05": "Mayıs", "06": "Haziran",
        "07": "Temmuz", "08": "Ağustos", "09": "Eylül", "10": "Ekim", "11": "Kasım", "12": "Aralık"
    }
    
    # Initialize counts
    summary_data = {m_code: {"count": 0, "base": 0, "kdv": 0, "stamp": 0, "turkak": 0, "total": 0} for m_code in months_tr}
    
    for p in protocols:
        m_part = p.payment_date.split("-")[1]
        if m_part in summary_data:
            summary_data[m_part]["count"] += 1
            summary_data[m_part]["base"] += (p.base_cost or 0)
            summary_data[m_part]["kdv"] += (p.kdv_amount or 0)
            summary_data[m_part]["stamp"] += (p.stamp_tax or 0)
            summary_data[m_part]["turkak"] += (p.turkak_fee or 0)
            summary_data[m_part]["total"] += (p.total_amount or 0)

    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{year} Yılı Deney Ücretleri"

    # Styles
    title_font = Font(name='Calibri', bold=True, size=11)
    header_font = Font(name='Calibri', bold=True, size=10)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border_style = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    fill_header = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")

    # Header section
    ws.merge_cells('A1:G1')
    ws['A1'] = "Karayolları 9. Bölge Müdürlüğü"
    ws.merge_cells('A2:G2')
    ws['A2'] = "(Araştırma ve Geliştirme Başmühendisliği)"
    ws.merge_cells('A3:G3')
    ws['A3'] = f"01 Ocak {year}-31 Aralık {year} Yılı Deney Ücretleri"
    for r in range(1, 4):
        ws.cell(row=r, column=1).alignment = center_align
        ws.cell(row=r, column=1).font = title_font

    # Column Headers
    headers = ["Ay", "Protokol Adedi", "Protokollü Hizmet Bedeli", "KDV", "Damga Vergisi", "Türkak Payı", "Genel Toplam"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border_style
        cell.fill = fill_header

    # Data Rows
    current_row = 6
    grand_totals = {"count": 0, "base": 0, "kdv": 0, "stamp": 0, "turkak": 0, "total": 0}
    for m_code in sorted(summary_data.keys()):
        data = summary_data[m_code]
        ws.cell(row=current_row, column=1, value=months_tr[m_code]).border = border_style
        ws.cell(row=current_row, column=2, value=data["count"]).border = border_style
        ws.cell(row=current_row, column=3, value=f"{data['base']:.2f} ₺").border = border_style
        ws.cell(row=current_row, column=4, value=f"{data['kdv']:.2f} ₺").border = border_style
        ws.cell(row=current_row, column=5, value=f"{data['stamp']:.2f} ₺").border = border_style
        ws.cell(row=current_row, column=6, value=f"{data['turkak']:.2f} ₺").border = border_style
        ws.cell(row=current_row, column=7, value=f"{data['total']:.2f} ₺").border = border_style
        
        grand_totals["count"] += data["count"]
        grand_totals["base"] += data["base"]
        grand_totals["kdv"] += data["kdv"]
        grand_totals["stamp"] += data["stamp"]
        grand_totals["turkak"] += data["turkak"]
        grand_totals["total"] += data["total"]
        current_row += 1

    # Totals Row
    ws.cell(row=current_row, column=1, value="Toplamlar").border = border_style
    ws.cell(row=current_row, column=1).font = header_font
    ws.cell(row=current_row, column=1).fill = fill_header
    
    ws.cell(row=current_row, column=2, value=grand_totals["count"]).border = border_style
    ws.cell(row=current_row, column=3, value=f"{grand_totals['base']:.2f} ₺").border = border_style
    ws.cell(row=current_row, column=4, value=f"{grand_totals['kdv']:.2f} ₺").border = border_style
    ws.cell(row=current_row, column=5, value=f"{grand_totals['stamp']:.2f} ₺").border = border_style
    ws.cell(row=current_row, column=6, value=f"{grand_totals['turkak']:.2f} ₺").border = border_style
    ws.cell(row=current_row, column=7, value=f"{grand_totals['total']:.2f} ₺").border = border_style
    for c in range(1, 8):
        ws.cell(row=current_row, column=c).font = header_font
        ws.cell(row=current_row, column=c).fill = fill_header

    # Signatures
    current_row += 2
    ws.merge_cells(f'A{current_row}:B{current_row}')
    ws[f'A{current_row}'] = "Hazırlayan"
    ws.merge_cells(f'C{current_row}:E{current_row}')
    ws[f'C{current_row}'] = "Kontrol Eden"
    ws.merge_cells(f'F{current_row}:G{current_row}')
    ws[f'F{current_row}'] = "Onaylayan"
    
    current_row += 1
    ws.merge_cells(f'A{current_row}:B{current_row}')
    ws[f'A{current_row}'] = "Sibel GÜMÜŞ"
    ws.merge_cells(f'C{current_row}:E{current_row}')
    ws[f'C{current_row}'] = "Jiyan KARAKAŞ"
    ws.merge_cells(f'F{current_row}:G{current_row}')
    ws[f'F{current_row}'] = "Atakan ERSOY"

    current_row += 1
    ws.merge_cells(f'A{current_row}:B{current_row}')
    ws[f'A{current_row}'] = "Büro ve Kayıt Memuru"
    ws.merge_cells(f'C{current_row}:E{current_row}')
    ws[f'C{current_row}'] = "Üstyapı Gel.Müh."
    ws.merge_cells(f'F{current_row}:G{current_row}')
    ws[f'F{current_row}'] = "Ar-Ge Başmühendisi"

    # Set column widths
    ws.column_dimensions['A'].width = 15
    for c in "BCDEFG":
        ws.column_dimensions[c].width = 20

    wb.save(output)
    output.seek(0)
    
    headers = {
        'Content-Disposition': f'attachment; filename="Yillik_Detayli_Ozet_{year}.xlsx"'
    }
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)

@app.get("/api/reports/annual-simplified/{year}")
def export_annual_simplified(year: str, db: Session = Depends(get_db)):
    # Resim 2 Formatı: AY, PROTOKOL SAYISI, PARA MİKTARI (TL), NOT
    protocols = db.query(Protocol).filter(Protocol.payment_date.like(f"{year}-%")).all()
    
    months_tr = {
        "01": "OCAK", "02": "ŞUBAT", "03": "MART", "04": "NİSAN", "05": "MAYIS", "06": "HAZİRAN",
        "07": "TEMMUZ", "08": "AĞUSTOS", "09": "EYLÜL", "10": "EKİM", "11": "KASIM", "12": "ARALIK"
    }
    
    summary_data = {m_code: {"count": 0, "total": 0} for m_code in months_tr}
    for p in protocols:
        m_part = p.payment_date.split("-")[1]
        if m_part in summary_data:
            summary_data[m_part]["count"] += 1
            summary_data[m_part]["total"] += (p.total_amount or 0)

    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{year} Protokol Listesi"

    header_font = Font(name='Calibri', bold=True, size=11)
    center_align = Alignment(horizontal='center', vertical='center')
    border_style = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Title section
    ws.merge_cells('A1:D1')
    ws['A1'] = f"01.01-31.12.{year} TARİHLERİ ARASINDA"
    ws.merge_cells('A2:D2')
    ws['A2'] = "KARAYOLLARI 9. BÖLGE MÜDÜRLÜĞÜ"
    ws.merge_cells('A3:D3')
    ws['A3'] = "ARAŞTIRMA VE GELİŞTİRME BAŞMÜHENDİSLİĞİ LAB.YAPILAN DENEY"
    ws.merge_cells('A4:D4')
    ws['A4'] = "PROTOKOL LİSTESİ"
    for r in range(1, 5):
        ws.cell(row=r, column=1).alignment = center_align
        ws.cell(row=r, column=1).font = header_font

    # Column Headers
    headers = ["AY", "PROTOKOL SAYISI", "PARA MİKTARI (TL)", "NOT"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border_style

    # Data
    current_row = 7
    total_count = 0
    total_amount = 0
    for m_code in sorted(summary_data.keys()):
        data = summary_data[m_code]
        ws.cell(row=current_row, column=1, value=months_tr[m_code]).border = border_style
        ws.cell(row=current_row, column=2, value=data["count"]).border = border_style
        ws.cell(row=current_row, column=3, value=f"{data['total']:.2f}").border = border_style
        ws.cell(row=current_row, column=4, value="").border = border_style
        
        total_count += data["count"]
        total_amount += data["total"]
        current_row += 1

    # Totals Row
    ws.cell(row=current_row, column=1, value="TOPLAM").border = border_style
    ws.cell(row=current_row, column=1).font = header_font
    ws.cell(row=current_row, column=2, value=total_count).border = border_style
    ws.cell(row=current_row, column=3, value=f"{total_amount:.2f}").border = border_style
    ws.cell(row=current_row, column=4, value="").border = border_style

    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 30

    wb.save(output)
    output.seek(0)
    
    headers = {
        'Content-Disposition': f'attachment; filename="Yillik_Basit_Ozet_{year}.xlsx"'
    }
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)

@app.get("/api/protocols/export")
def export_protocols(month: Optional[str] = None, db: Session = Depends(get_db)):
    try:
        query = db.query(Protocol).filter(Protocol.is_archived == 0)
        month_name = "TÜM ZAMANLAR"
        if month:
            query = query.filter(Protocol.payment_date.like(f"{month}-%"))
            # Get month name for title
            months_tr = {"01":"OCAK","02":"ŞUBAT","03":"MART","04":"NİSAN","05":"MAYIS","06":"HAZİRAN","07":"TEMMUZ","08":"AĞUSTOS","09":"EYLÜL","10":"EKİM","11":"KASIM","12":"ARALIK"}
            m_part = month.split("-")[1]
            year_part = month.split("-")[0]
            month_name = f"{year_part} YILI {months_tr.get(m_part, '')} AYI"
        
        protocols = query.order_by(Protocol.payment_date.asc()).all()
        
        output = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Protokol Listesi"

        # Styles
        header_font = Font(bold=True, size=12)
        yellow_fill = PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        bold_font = Font(bold=True)

        # Upper Header
        ws.merge_cells('A1:R1')
        ws['A1'] = "PROTOKOL LİSTESİ"
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws['A1'].font = Font(bold=True, size=14)

        ws.merge_cells('A2:R2')
        ws['A2'] = "Karayolları 9. Bölge Müdürlüğü"
        ws['A2'].alignment = Alignment(horizontal="center", vertical="center")
        ws['A2'].font = Font(bold=True, size=12)

        ws.merge_cells('A3:R3')
        ws['A3'] = month_name
        ws['A3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['A3'].font = Font(bold=True, size=12)

        ws.merge_cells('A4:R4')
        ws['A4'] = "AR-GE Dairesi Bşk. Hizmetleri Protokollü İşler Listesi"
        ws['A4'].alignment = Alignment(horizontal="center", vertical="center")
        ws['A4'].font = Font(bold=True, size=11)

        # Table Headers (Row 6 & 7)
        headers = [
            "Sıra No:", "Büro Kayıt No:", "Protokol No:", "Bölge No", "NUMUNEYİ GÖNDEREN", 
            "PROTOKOL BEDELİNİ YATIRAN FİRMA", "YAPILACAK İŞ", "Protokol İmzalanma Tarihi",
            "Deney 1.Keşif Özeti (TL)", "KDV ( H * %20 ) (TL)", "KDV' li 1.Keşif Özeti (TL)",
            "KDV' li 2.Keşif Tutarı ( TL )", "TÜRKAK Payı ( G* %06 )", "Damga Vergisi ( G*0,00948 )",
            "Toplam Miktar ( TL )", "DEKONT Tarihi", "DEKONT NO:", "AÇIKLAMA"
        ]
        
        sub_headers = [
            "", "", "", "", "", "", "", "", "H", "I", "H+I", "i", "K", "", "H+K+I+(i)", "", "", ""
        ]

        for col, text in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col)
            cell.value = text
            cell.fill = yellow_fill
            cell.font = bold_font
            cell.border = border
            cell.alignment = center_align

        for col, text in enumerate(sub_headers, 1):
            cell = ws.cell(row=7, column=col)
            cell.value = text
            cell.fill = yellow_fill
            cell.font = bold_font
            cell.border = border
            cell.alignment = center_align

        ws.row_dimensions[6].height = 60

        # Data Rows
        row_idx = 8
        for p in protocols:
            ws.cell(row=row_idx, column=1, value=p.sequence_no).border = border
            ws.cell(row=row_idx, column=2, value=p.office_record_no).border = border
            ws.cell(row=row_idx, column=3, value=p.protocol_no).border = border
            ws.cell(row=row_idx, column=4, value=p.region_no).border = border
            ws.cell(row=row_idx, column=5, value=p.sender).border = border
            ws.cell(row=row_idx, column=6, value=p.firm).border = border
            ws.cell(row=row_idx, column=7, value=p.job_description).border = border
            ws.cell(row=row_idx, column=8, value=p.protocol_date).border = border
            ws.cell(row=row_idx, column=9, value=p.base_cost).border = border
            ws.cell(row=row_idx, column=10, value=p.kdv_amount).border = border
            ws.cell(row=row_idx, column=11, value=(p.base_cost or 0) + (p.kdv_amount or 0)).border = border
            ws.cell(row=row_idx, column=12, value=p.secondary_keşif_with_kdv).border = border
            ws.cell(row=row_idx, column=13, value=p.turkak_fee).border = border
            ws.cell(row=row_idx, column=14, value=p.stamp_tax).border = border
            ws.cell(row=row_idx, column=15, value=p.total_amount).border = border
            ws.cell(row=row_idx, column=16, value=p.payment_date).border = border
            ws.cell(row=row_idx, column=17, value=p.receipt_no).border = border
            ws.cell(row=row_idx, column=18, value=p.bank_info).border = border
            
            # Numeric formatting for financial columns
            for col in range(9, 16):
                ws.cell(row=row_idx, column=col).number_format = '#,##0.00'
            
            # Center align everything
            for col in range(1, 19):
                ws.cell(row=row_idx, column=col).alignment = center_align
            
            row_idx += 1

        # Total Row
        ws.merge_cells(f'A{row_idx}:H{row_idx}')
        total_label = ws.cell(row=row_idx, column=1)
        total_label.value = "TOPLAM"
        total_label.alignment = Alignment(horizontal="right")
        total_label.font = bold_font
        total_label.fill = yellow_fill
        total_label.border = border

        for col in range(9, 16):
            col_letter = get_column_letter(col)
            ws.cell(row=row_idx, column=col, value=f"=SUM({col_letter}8:{col_letter}{row_idx-1})").border = border
            ws.cell(row=row_idx, column=col).font = bold_font
            ws.cell(row=row_idx, column=col).alignment = center_align

        # Footer
        row_idx += 2
        ws.cell(row=row_idx, column=1, value="NOT 1: H, I, I, G, J Keşif Özetinde Verilen Miktarlar Olup Keşif Özetinden Alınacaktır.")
        row_idx += 2
        ws.cell(row=row_idx, column=1, value="NOT 2: Toplam Miktar hesaplanırken ( i ) veya ( I )'den Biri Alınacak. ")

        # Signature Block
        row_idx += 1
        # Right-aligned block
        ws.cell(row=row_idx, column=10, value="Hazırlayan").font = bold_font
        ws.cell(row=row_idx, column=13, value="................................")
        row_idx += 1
        ws.cell(row=row_idx, column=10, value="Adı-Soyadı").font = bold_font
        ws.cell(row=row_idx, column=13, value="................................")
        row_idx += 1
        ws.cell(row=row_idx, column=10, value="Ünvanı").font = bold_font
        ws.cell(row=row_idx, column=13, value="................................")
        row_idx += 1
        ws.cell(row=row_idx, column=10, value="İmza").font = bold_font
        ws.cell(row=row_idx, column=13, value="................................")

        # Column Widths
        ws.column_dimensions['A'].width = 8   # Sıra No
        ws.column_dimensions['B'].width = 15  # Büro Kayıt No
        ws.column_dimensions['C'].width = 12  # Protokol No
        ws.column_dimensions['D'].width = 10  # Bölge No
        ws.column_dimensions['E'].width = 30  # NUMUNEYİ GÖNDEREN
        ws.column_dimensions['F'].width = 30  # FİRMA
        ws.column_dimensions['G'].width = 35  # İŞ TANIMI
        ws.column_dimensions['H'].width = 20  # TARİH
        
        for col in range(9, 16):
            ws.column_dimensions[get_column_letter(col)].width = 18

        ws.column_dimensions['P'].width = 15  # Dekont Tarih
        ws.column_dimensions['Q'].width = 15  # Dekont No
        ws.column_dimensions['R'].width = 25  # Açıklama

        wb.save(output)
        output.seek(0)
        
        filename = f"protokol_listesi_{month if month else 'tum'}.xlsx"
        headers = {
            'Content-Disposition': f'attachment; filename="{filename}"'
        }
        return StreamingResponse(output, headers=headers, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        import traceback
        with open("debug_error.log", "a") as f:
            f.write(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/protocols/{id}/pdf")
def get_protocol_pdf(id: int, db: Session = Depends(get_db)):
    p = db.query(Protocol).filter(Protocol.id == id).first()
    if p is None:
        raise HTTPException(status_code=404, detail="Protocol not found")
    
    pdf = FPDF()
    pdf.add_page()
    
    def tr(text):
        if not text: return ""
        replacements = {'ı': 'i', 'İ': 'I', 'ğ': 'g', 'Ğ': 'G', 'ü': 'u', 'Ü': 'U', 'ş': 's', 'Ş': 'S', 'ö': 'o', 'Ö': 'O', 'ç': 'c', 'Ç': 'C'}
        for k, v in replacements.items():
            text = text.replace(k, v)
        return text

    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, tr("PROTOKOL RAPORU"), ln=True, align="C")
    pdf.ln(10)
    
    pdf.set_font("Helvetica", size=12)
    fields = [
        ("Protokol No", p.protocol_no),
        ("Firma", p.firm),
        ("Is Tanimi", p.job_description),
        ("Tarih", p.protocol_date),
        ("Toplam Tutar", f"{p.total_amount} TL"),
        ("Banka Bilgisi", p.bank_info),
        ("Ay", p.month)
    ]
    
    for label, value in fields:
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(40, 10, tr(label) + ": ", 0)
        pdf.set_font("Helvetica", size=12)
        pdf.cell(0, 10, tr(str(value)), 1)
        pdf.ln(12)

    output = io.BytesIO(pdf.output())
    headers = {
        'Content-Disposition': f'attachment; filename="protokol_{id}.pdf"'
    }
    return StreamingResponse(output, headers=headers, media_type="application/pdf")

@app.post("/api/protocols", response_model=ProtocolSchema)
def create_protocol(protocol: ProtocolBase, db: Session = Depends(get_db)):
    db_protocol = Protocol(**protocol.dict())
    db.add(db_protocol)
    db.commit()
    db.refresh(db_protocol)
    return db_protocol

@app.get("/api/protocols/{id}", response_model=ProtocolSchema)
def read_protocol(id: int, db: Session = Depends(get_db)):
    db_protocol = db.query(Protocol).filter(Protocol.id == id).first()
    if db_protocol is None:
        raise HTTPException(status_code=404, detail="Protocol not found")
    return db_protocol

@app.put("/api/protocols/{id}", response_model=ProtocolSchema)
def update_protocol(id: int, protocol: ProtocolBase, db: Session = Depends(get_db)):
    db_protocol = db.query(Protocol).filter(Protocol.id == id).first()
    if db_protocol is None:
        raise HTTPException(status_code=404, detail="Protocol not found")
    
    # Use exclude_unset=True for partial updates
    update_data = protocol.dict(exclude_unset=True)
    for key, value in update_data.items():
        setattr(db_protocol, key, value)
    
    db.commit()
    db.refresh(db_protocol)
    return db_protocol

@app.delete("/api/protocols/{id}")
def delete_protocol(id: int, db: Session = Depends(get_db)):
    db_protocol = db.query(Protocol).filter(Protocol.id == id).first()
    if db_protocol is None:
        raise HTTPException(status_code=404, detail="Protocol not found")
    db.delete(db_protocol)
    db.commit()
    return {"detail": "Protocol deleted"}

@app.get("/api/stats")
def get_stats(db: Session = Depends(get_db)):
    # Only count active (non-archived) protocols
    total_count = db.query(Protocol).filter(Protocol.is_archived == 0).count()
    total_revenue = db.query(Protocol).with_entities(Protocol.total_amount).filter(Protocol.is_archived == 0).all()
    revenue_sum = sum(float(r[0]) for r in total_revenue if r[0])
    
    # Enhanced monthly breakdown for Chart.js based on payment_date (dekont tarihi)
    months_tr = ["Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran", "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık"]
    
    # Only chart active protocols
    query = db.query(Protocol.payment_date).filter(Protocol.is_archived == 0).all()
    month_counts = {m: 0 for m in months_tr}
    
    for (d_date,) in query:
        if d_date and "-" in d_date:
            try:
                # Format is YYYY-MM-DD
                month_idx = int(d_date.split("-")[1]) - 1
                if 0 <= month_idx < 12:
                    m_name = months_tr[month_idx]
                    month_counts[m_name] += 1
            except (ValueError, IndexError):
                continue
    
    # Only return months that have data to keep the chart clean, but maintain order
    labels = [m for m in months_tr if month_counts[m] > 0]
    values = [month_counts[m] for m in labels]
            
    return {
        "total_count": total_count,
        "total_revenue": revenue_sum,
        "chart_data": {
            "labels": labels,
            "values": values
        }
    }

# Serve static files
# Create public directory if it doesn't exist
if not os.path.exists("public"):
    os.makedirs("public")

app.mount("/static", StaticFiles(directory="public"), name="static")

@app.get("/")
def read_index():
    return FileResponse("public/index.html")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001)
